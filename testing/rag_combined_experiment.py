import json
import os
import site
import sys
from datetime import datetime
from pathlib import Path
from statistics import mean
from typing import Any, Dict, List

try:
    import openpyxl
except ModuleNotFoundError:
    candidate_site_packages = [
        *site.getsitepackages(),
        str(Path(sys.base_prefix) / "Lib" / "site-packages"),
    ]
    for site_packages in candidate_site_packages:
        if site_packages not in sys.path:
            sys.path.append(site_packages)
    import openpyxl


ROOT_DIR = Path(__file__).resolve().parents[1]
BACKEND_DIR = ROOT_DIR / "backend"
BASELINE_WORKBOOK = ROOT_DIR / "testing" / "AI_Mentor_Baseline_Chatbot_Testing.xlsx"
OUTPUT_XLSX = ROOT_DIR / "testing" / "rag_combined_experiment_results.xlsx"
OUTPUT_JSON = ROOT_DIR / "testing" / "rag_combined_experiment_results.json"

COURSE_ID = "CPL"
TOP_K = 5
DISTANCE_THRESHOLD = 1.4

MODE_MAP = {
    "Explain Assignment": "explain_assignment",
    "Explain Rubric": "explain_rubric",
    "Explain Topic": "explain_topic",
    "Review Draft": "review_draft",
    "Ask Anything": "ask_anything",
}


def load_backend_env() -> None:
    env_path = BACKEND_DIR / ".env"
    if not env_path.exists():
        return

    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


def mode_to_internal(label: str) -> str:
    if label not in MODE_MAP:
        raise ValueError(f"Unsupported baseline mentor mode: {label}")
    return MODE_MAP[label]


def load_baseline_cases() -> List[Dict[str, Any]]:
    workbook = openpyxl.load_workbook(BASELINE_WORKBOOK, data_only=True)
    worksheet = workbook["Baseline Tests"]
    cases = []

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        test_id = row[0]
        if not test_id:
            continue
        cases.append(
            {
                "test_id": test_id,
                "question_type": row[1],
                "cpl_area": row[2],
                "question": row[3],
                "expected_source_topic": row[4],
                "mentor_mode_label": row[5],
                "mentor_mode": mode_to_internal(row[5]),
                "baseline_note": row[19],
            }
        )

    if len(cases) != 15:
        raise RuntimeError(f"Expected 15 baseline cases, found {len(cases)}")

    return cases


def rank_results(results: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    ranked_results = []
    for rank, item in enumerate(results, start=1):
        distance = float(item.get("distance"))
        ranked_results.append(
            {
                "rank": rank,
                "chunk_id": item.get("chunk_id"),
                "title": item.get("title"),
                "instructional_unit": item.get("instructional_unit"),
                "document_type": item.get("document_type"),
                "source_file": item.get("source_file"),
                "distance": distance,
                "threshold_decision": "retained"
                if distance <= DISTANCE_THRESHOLD
                else "discarded",
            }
        )
    return ranked_results


def retrieve_cases(cases: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    sys.path.insert(0, str(BACKEND_DIR))
    load_backend_env()

    from app.services.retrieval_service import retrieve_cpl_context

    evaluated = []
    for case in cases:
        top5_results = rank_results(
            retrieve_cpl_context(
                query=case["question"],
                course_id=COURSE_ID,
                mentor_mode=case["mentor_mode"],
                top_k=TOP_K,
            )
        )
        retained = [
            item for item in top5_results if item["threshold_decision"] == "retained"
        ]
        discarded = [
            item for item in top5_results if item["threshold_decision"] == "discarded"
        ]

        evaluated.append(
            {
                **case,
                "top_k": TOP_K,
                "distance_threshold": DISTANCE_THRESHOLD,
                "original_count": len(top5_results),
                "remaining_count": len(retained),
                "context_status": "NO_CONTEXT" if not retained else "CONTEXT",
                "retrieved_chunks": top5_results,
                "retained_chunks": retained,
                "discarded_chunks": discarded,
            }
        )

    return evaluated


def summarize(evaluated: List[Dict[str, Any]]) -> Dict[str, Any]:
    before_counts = [case["original_count"] for case in evaluated]
    after_counts = [case["remaining_count"] for case in evaluated]
    removed_cases = [case for case in evaluated if case["discarded_chunks"]]

    return {
        "question_count": len(evaluated),
        "baseline": {
            "configuration": "Top 5 + no threshold",
            "average_chunks_before_filtering": mean(before_counts),
        },
        "experiment": {
            "configuration": "Top 5 + distance <= 1.4",
            "average_chunks_after_filtering": mean(after_counts),
            "tests_where_all_5_remain": [
                case["test_id"] for case in evaluated if case["remaining_count"] == 5
            ],
            "tests_where_1_or_more_chunks_removed": [
                case["test_id"] for case in removed_cases
            ],
            "tests_producing_no_context": [
                case["test_id"]
                for case in evaluated
                if case["context_status"] == "NO_CONTEXT"
            ],
            "removed_chunks_by_test": [
                {
                    "test_id": case["test_id"],
                    "removed_chunks": [
                        {
                            "rank": item["rank"],
                            "chunk_id": item["chunk_id"],
                            "title": item["title"],
                            "instructional_unit": item["instructional_unit"],
                            "document_type": item["document_type"],
                            "distance": item["distance"],
                        }
                        for item in case["discarded_chunks"]
                    ],
                }
                for case in removed_cases
            ],
            "t14_produces_no_context": next(
                case for case in evaluated if case["test_id"] == "T14"
            )["context_status"]
            == "NO_CONTEXT",
            "t04": {
                "remaining_count": next(
                    case for case in evaluated if case["test_id"] == "T04"
                )["remaining_count"],
                "context_status": next(
                    case for case in evaluated if case["test_id"] == "T04"
                )["context_status"],
                "removed_chunks": next(
                    case for case in evaluated if case["test_id"] == "T04"
                )["discarded_chunks"],
            },
            "t15": {
                "remaining_count": next(
                    case for case in evaluated if case["test_id"] == "T15"
                )["remaining_count"],
                "context_status": next(
                    case for case in evaluated if case["test_id"] == "T15"
                )["context_status"],
                "removed_chunks": next(
                    case for case in evaluated if case["test_id"] == "T15"
                )["discarded_chunks"],
            },
        },
    }


def write_workbook(evaluated: List[Dict[str, Any]], summary: Dict[str, Any]) -> None:
    workbook = openpyxl.Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Combined Summary"
    summary_sheet.append(["Metric", "Value"])
    summary_sheet.append(["Baseline", summary["baseline"]["configuration"]])
    summary_sheet.append(["Experiment", summary["experiment"]["configuration"]])
    summary_sheet.append(
        [
            "Average chunks before filtering",
            round(summary["baseline"]["average_chunks_before_filtering"], 2),
        ]
    )
    summary_sheet.append(
        [
            "Average chunks after filtering",
            round(summary["experiment"]["average_chunks_after_filtering"], 2),
        ]
    )
    summary_sheet.append(
        [
            "Tests where all 5 remain",
            ", ".join(summary["experiment"]["tests_where_all_5_remain"]) or "None",
        ]
    )
    summary_sheet.append(
        [
            "Tests where 1+ chunks removed",
            ", ".join(summary["experiment"]["tests_where_1_or_more_chunks_removed"])
            or "None",
        ]
    )
    summary_sheet.append(
        [
            "Tests producing NO_CONTEXT",
            ", ".join(summary["experiment"]["tests_producing_no_context"]) or "None",
        ]
    )
    summary_sheet.append(
        [
            "T14 produces NO_CONTEXT",
            "Yes" if summary["experiment"]["t14_produces_no_context"] else "No",
        ]
    )

    detail_sheet = workbook.create_sheet("Retrieved Chunks")
    detail_sheet.append(
        [
            "Test ID",
            "Question",
            "Mentor Mode",
            "Original Rank",
            "Chunk ID",
            "Title",
            "Instructional Unit",
            "Document Type",
            "Distance",
            "Threshold Decision",
            "Remaining Count",
            "Context Status",
        ]
    )
    for case in evaluated:
        for item in case["retrieved_chunks"]:
            detail_sheet.append(
                [
                    case["test_id"],
                    case["question"],
                    case["mentor_mode"],
                    item["rank"],
                    item["chunk_id"],
                    item["title"],
                    item["instructional_unit"],
                    item["document_type"],
                    round(item["distance"], 6),
                    item["threshold_decision"],
                    case["remaining_count"],
                    case["context_status"],
                ]
            )

    removed_sheet = workbook.create_sheet("Removed Chunks")
    removed_sheet.append(
        [
            "Test ID",
            "Original Rank",
            "Chunk ID",
            "Title",
            "Instructional Unit",
            "Document Type",
            "Distance",
        ]
    )
    for case in evaluated:
        for item in case["discarded_chunks"]:
            removed_sheet.append(
                [
                    case["test_id"],
                    item["rank"],
                    item["chunk_id"],
                    item["title"],
                    item["instructional_unit"],
                    item["document_type"],
                    round(item["distance"], 6),
                ]
            )

    workbook.save(OUTPUT_XLSX)


def write_json(evaluated: List[Dict[str, Any]], summary: Dict[str, Any]) -> None:
    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "course_id": COURSE_ID,
        "top_k": TOP_K,
        "distance_threshold": DISTANCE_THRESHOLD,
        "threshold_rule": "retain when distance <= 1.4",
        "llm_called": False,
        "relevance_classification": "none",
        "embedding_model": "text-embedding-3-small",
        "chroma_collection": "cpl_documents",
        "summary": summary,
        "cases": evaluated,
    }
    OUTPUT_JSON.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def main() -> None:
    cases = load_baseline_cases()
    evaluated = retrieve_cases(cases)
    summary = summarize(evaluated)

    write_workbook(evaluated, summary)
    write_json(evaluated, summary)

    print(f"Wrote {OUTPUT_XLSX}")
    print(f"Wrote {OUTPUT_JSON}")
    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
