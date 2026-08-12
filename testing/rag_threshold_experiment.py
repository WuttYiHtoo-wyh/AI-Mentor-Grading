import json
import os
import re
import sys
from datetime import datetime
from pathlib import Path
from statistics import mean
from typing import Any, Dict, List

import openpyxl


ROOT_DIR = Path(__file__).resolve().parents[1]
BACKEND_DIR = ROOT_DIR / "backend"
BASELINE_WORKBOOK = ROOT_DIR / "testing" / "AI_Mentor_Baseline_Chatbot_Testing.xlsx"
OUTPUT_XLSX = ROOT_DIR / "testing" / "rag_threshold_experiment_results.xlsx"
OUTPUT_JSON = ROOT_DIR / "testing" / "rag_threshold_experiment_results.json"

COURSE_ID = "CPL"
DEFAULT_TOP_K = 5
CANDIDATE_THRESHOLDS = [1.2, 1.3, 1.4, 1.5]

MODE_MAP = {
    "Explain Assignment": "explain_assignment",
    "Explain Rubric": "explain_rubric",
    "Explain Topic": "explain_topic",
    "Review Draft": "review_draft",
    "Ask Anything": "ask_anything",
}

# These indicators are used only to highlight threshold risk in this experiment.
# They do not change retrieval, scoring, prompts, or production configuration.
POTENTIAL_RELEVANCE_INDICATORS = {
    "T01": ["gantt"],
    "T02": ["gantt", "work breakdown", "wbs", "dependency", "dependencies", "schedule"],
    "T03": ["research question", "too broad", "scope", "specific", "literature review"],
    "T04": ["sample", "sampling", "participant", "survey", "methodology", "limitation", "response"],
    "T05": ["research topic", "topic selection", "scope", "feasible", "interest"],
    "T06": ["literature review", "source", "gap", "background", "existing research"],
    "T07": ["qualitative", "quantitative", "methodology", "experience", "detail"],
    "T08": ["primary research", "secondary research", "survey", "websites", "reports"],
    "T09": ["data analysis", "prepare data", "clean", "survey data", "findings"],
    "T10": ["finding", "recommendation"],
    "T11": ["reflection", "future research", "limitation", "evaluate"],
    "T12": ["alignment", "research question", "survey question", "finding", "recommendation"],
    "T13": ["project planning", "planning", "wbs", "gantt", "schedule", "milestone", "risk"],
    "T14": [],
    "T15": ["survey", "participant", "sample", "respondent", "questionnaire", "data collection"],
}


def load_backend_env() -> None:
    env_path = BACKEND_DIR / ".env"
    if not env_path.exists():
        return

    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


def normalize_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip().lower()


def mode_to_internal(label: str) -> str:
    if label not in MODE_MAP:
        raise ValueError(f"Unsupported baseline mentor mode: {label}")
    return MODE_MAP[label]


def load_baseline_cases() -> List[Dict[str, Any]]:
    workbook = openpyxl.load_workbook(BASELINE_WORKBOOK, data_only=True)
    worksheet = workbook["Baseline Tests"]
    cases = []

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        test_id = row[0]
        if not test_id:
            continue
        cases.append(
            {
                "test_id": test_id,
                "question_type": row[1],
                "cpl_area": row[2],
                "question": row[3],
                "expected_source_topic": row[4],
                "mentor_mode_label": row[5],
                "mentor_mode": mode_to_internal(row[5]),
                "baseline_note": row[19],
            }
        )

    if len(cases) != 15:
        raise RuntimeError(f"Expected 15 baseline cases, found {len(cases)}")

    return cases


def is_potentially_relevant(test_id: str, item: Dict[str, Any]) -> bool:
    indicators = POTENTIAL_RELEVANCE_INDICATORS.get(test_id, [])
    if not indicators:
        return False

    metadata = item.get("metadata") or {}
    haystack = normalize_text(
        " ".join(
            str(value or "")
            for value in [
                item.get("title"),
                item.get("instructional_unit"),
                item.get("document_type"),
                item.get("source_file"),
                item.get("content"),
                metadata.get("keywords"),
                metadata.get("question"),
                metadata.get("topic"),
                metadata.get("section"),
            ]
        )
    )
    return any(indicator in haystack for indicator in indicators)


def retrieve_cases(cases: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    sys.path.insert(0, str(BACKEND_DIR))
    load_backend_env()

    from app.services.retrieval_service import retrieve_cpl_context

    evaluated = []
    for case in cases:
        results = retrieve_cpl_context(
            query=case["question"],
            course_id=COURSE_ID,
            mentor_mode=case["mentor_mode"],
            top_k=DEFAULT_TOP_K,
        )

        ranked_results = []
        for rank, item in enumerate(results, start=1):
            distance = float(item.get("distance"))
            ranked_results.append(
                {
                    "rank": rank,
                    "chunk_id": item.get("chunk_id"),
                    "title": item.get("title"),
                    "instructional_unit": item.get("instructional_unit"),
                    "document_type": item.get("document_type"),
                    "source_file": item.get("source_file"),
                    "distance": distance,
                    "potentially_relevant": is_potentially_relevant(case["test_id"], item),
                }
            )

        threshold_outcomes = {}
        for threshold in CANDIDATE_THRESHOLDS:
            kept = [item for item in ranked_results if item["distance"] <= threshold]
            discarded = [item for item in ranked_results if item["distance"] > threshold]
            threshold_outcomes[str(threshold)] = {
                "kept_ranks": [item["rank"] for item in kept],
                "discarded_ranks": [item["rank"] for item in discarded],
                "kept_chunk_ids": [item["chunk_id"] for item in kept],
                "discarded_chunk_ids": [item["chunk_id"] for item in discarded],
                "retained_count": len(kept),
                "all_rejected": len(kept) == 0,
                "lost_potentially_relevant": [
                    {
                        "rank": item["rank"],
                        "chunk_id": item["chunk_id"],
                        "title": item["title"],
                        "distance": item["distance"],
                    }
                    for item in discarded
                    if item["potentially_relevant"]
                ],
            }

        evaluated.append({**case, "results": ranked_results, "thresholds": threshold_outcomes})

    return evaluated


def summarize(evaluated: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    summary = []
    for threshold in CANDIDATE_THRESHOLDS:
        key = str(threshold)
        retained_counts = [case["thresholds"][key]["retained_count"] for case in evaluated]
        all_rejected = [
            case["test_id"] for case in evaluated if case["thresholds"][key]["all_rejected"]
        ]
        lost_cases = []
        for case in evaluated:
            lost = case["thresholds"][key]["lost_potentially_relevant"]
            if lost:
                lost_cases.append(
                    {
                        "test_id": case["test_id"],
                        "lost_chunks": [
                            f"{item['chunk_id']} r{item['rank']} ({item['distance']:.4f})"
                            for item in lost
                        ],
                    }
                )

        t14 = next(case for case in evaluated if case["test_id"] == "T14")
        summary.append(
            {
                "threshold": threshold,
                "average_chunks_retained": mean(retained_counts),
                "questions_retaining_at_least_1": sum(1 for count in retained_counts if count > 0),
                "all_5_rejected_questions": all_rejected,
                "cases_losing_potentially_relevant_chunks": lost_cases,
                "t14_successfully_rejected": t14["thresholds"][key]["all_rejected"],
            }
        )
    return summary


def write_workbook(evaluated: List[Dict[str, Any]], summary: List[Dict[str, Any]]) -> None:
    workbook = openpyxl.Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Threshold Summary"
    summary_sheet.append(
        [
            "Threshold",
            "Average Chunks Retained",
            "Questions Retaining >=1 Chunk",
            "All 5 Rejected Questions",
            "Cases Losing Potentially Relevant Chunks",
            "T14 Successfully Rejected",
        ]
    )
    for row in summary:
        summary_sheet.append(
            [
                row["threshold"],
                round(row["average_chunks_retained"], 2),
                row["questions_retaining_at_least_1"],
                ", ".join(row["all_5_rejected_questions"]) or "None",
                "; ".join(
                    f"{case['test_id']}: {', '.join(case['lost_chunks'])}"
                    for case in row["cases_losing_potentially_relevant_chunks"]
                )
                or "None",
                "Yes" if row["t14_successfully_rejected"] else "No",
            ]
        )

    detail_sheet = workbook.create_sheet("Top 5 Detail")
    detail_sheet.append(
        [
            "Test ID",
            "Question",
            "Mentor Mode",
            "Rank",
            "Chunk ID",
            "Title",
            "Instructional Unit",
            "Document Type",
            "Distance",
            "Potentially Relevant",
            "1.2",
            "1.3",
            "1.4",
            "1.5",
        ]
    )
    for case in evaluated:
        for item in case["results"]:
            detail_sheet.append(
                [
                    case["test_id"],
                    case["question"],
                    case["mentor_mode"],
                    item["rank"],
                    item["chunk_id"],
                    item["title"],
                    item["instructional_unit"],
                    item["document_type"],
                    round(item["distance"], 6),
                    "Yes" if item["potentially_relevant"] else "No",
                    *[
                        "KEEP" if item["distance"] <= threshold else "DISCARD"
                        for threshold in CANDIDATE_THRESHOLDS
                    ],
                ]
            )

    lost_sheet = workbook.create_sheet("Potential Relevance Loss")
    lost_sheet.append(["Threshold", "Test ID", "Lost Chunk", "Rank", "Title", "Distance"])
    for row in summary:
        threshold = row["threshold"]
        threshold_key = str(threshold)
        for case in evaluated:
            for item in case["thresholds"][threshold_key]["lost_potentially_relevant"]:
                lost_sheet.append(
                    [
                        threshold,
                        case["test_id"],
                        item["chunk_id"],
                        item["rank"],
                        item["title"],
                        round(item["distance"], 6),
                    ]
                )

    workbook.save(OUTPUT_XLSX)


def write_json(evaluated: List[Dict[str, Any]], summary: List[Dict[str, Any]]) -> None:
    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "course_id": COURSE_ID,
        "top_k": DEFAULT_TOP_K,
        "thresholds": CANDIDATE_THRESHOLDS,
        "summary": summary,
        "cases": evaluated,
    }
    OUTPUT_JSON.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def main() -> None:
    cases = load_baseline_cases()
    evaluated = retrieve_cases(cases)
    summary = summarize(evaluated)

    write_workbook(evaluated, summary)
    write_json(evaluated, summary)

    print(f"Wrote {OUTPUT_XLSX}")
    print(f"Wrote {OUTPUT_JSON}")
    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
