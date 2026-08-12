import json
import os
import shutil
import site
import sys
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

try:
    import openpyxl
except ModuleNotFoundError:
    candidate_site_packages = [
        *site.getsitepackages(),
        str(Path(sys.base_prefix) / "Lib" / "site-packages"),
    ]
    for site_packages in candidate_site_packages:
        if site_packages not in sys.path:
            sys.path.append(site_packages)
    import openpyxl


ROOT_DIR = Path(__file__).resolve().parents[1]
BACKEND_DIR = ROOT_DIR / "backend"
CHROMA_DB_DIR = BACKEND_DIR / "app" / "chroma_db"
BASELINE_V1_JSON = ROOT_DIR / "testing" / "ai_mentor_behaviour_baseline.json"
BASELINE_V2_JSON = ROOT_DIR / "testing" / "ai_mentor_behaviour_v2_regression.json"
OUTPUT_XLSX = ROOT_DIR / "testing" / "ai_mentor_behaviour_v3_targeted.xlsx"
OUTPUT_JSON = ROOT_DIR / "testing" / "ai_mentor_behaviour_v3_targeted.json"

COURSE_ID = "CPL"
TARGET_IDS = ["P05", "P07", "P08", "P09", "P10"]


def load_backend_env() -> None:
    env_path = BACKEND_DIR / ".env"
    if not env_path.exists():
        return

    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


def load_cases() -> List[Dict[str, Any]]:
    payload = json.loads(BASELINE_V1_JSON.read_text(encoding="utf-8"))
    cases_by_id = {case["test_id"]: case for case in payload.get("cases", [])}
    missing = [test_id for test_id in TARGET_IDS if test_id not in cases_by_id]
    if missing:
        raise RuntimeError(f"Missing target cases in V1 baseline: {missing}")
    return [cases_by_id[test_id] for test_id in TARGET_IDS]


def load_v2_by_id() -> Dict[str, Dict[str, Any]]:
    if not BASELINE_V2_JSON.exists():
        return {}
    payload = json.loads(BASELINE_V2_JSON.read_text(encoding="utf-8"))
    return {case["test_id"]: case for case in payload.get("cases", [])}


def compact_retrieval_item(item: Dict[str, Any], rank: int) -> Dict[str, Any]:
    return {
        "rank": rank,
        "chunk_id": item.get("chunk_id"),
        "title": item.get("title"),
        "document_type": item.get("document_type"),
        "instructional_unit": item.get("instructional_unit"),
        "source_file": item.get("source_file"),
        "distance": float(item.get("distance")),
    }


def run_targeted_v3(cases: List[Dict[str, Any]], v2_by_id: Dict[str, Dict[str, Any]]) -> List[Dict[str, Any]]:
    sys.path.insert(0, str(BACKEND_DIR))
    load_backend_env()

    import chromadb
    from chromadb.config import Settings

    import app.services.retrieval_service as retrieval_service
    from app.services.conversation_router import get_conversational_response
    from app.services.llm_service import LLMService

    temp_dir = Path(tempfile.mkdtemp(prefix="ai_mentor_v3_targeted_chroma_"))
    temp_chroma_dir = temp_dir / "chroma_db"
    original_get_chroma_client = retrieval_service.get_chroma_client

    try:
        shutil.copytree(CHROMA_DB_DIR, temp_chroma_dir)
        temp_client = chromadb.Client(
            Settings(persist_directory=str(temp_chroma_dir), is_persistent=True)
        )
        retrieval_service.get_chroma_client = lambda: temp_client
        llm = LLMService()

        evaluated = []
        for baseline in cases:
            runtime_error: Optional[str] = None
            answer: Optional[str] = None
            sources: List[Dict[str, str]] = []
            conversation_id: Optional[str] = None
            retrieval_context: List[Dict[str, Any]] = []
            route = "retrieval_llm"

            try:
                conversational_response = get_conversational_response(
                    message=baseline["question"],
                    conversation_id=None,
                )
                if conversational_response:
                    route = "conversation_router"
                    answer = conversational_response.answer
                    sources = [
                        source.model_dump() for source in conversational_response.sources
                    ]
                    conversation_id = conversational_response.conversation_id
                else:
                    retrieval_context = retrieval_service.retrieve_cpl_context(
                        query=baseline["question"],
                        course_id=COURSE_ID,
                        mentor_mode=baseline["mentor_mode"],
                        top_k=retrieval_service.DEFAULT_TOP_K,
                    )
                    response = llm.generate_chat_response(
                        message=baseline["question"],
                        conversation_id=None,
                        mentor_mode=baseline["mentor_mode"],
                        retrieval_context=retrieval_context,
                    )
                    answer = response.get("answer")
                    sources = [
                        source.model_dump()
                        if hasattr(source, "model_dump")
                        else dict(source)
                        for source in response.get("sources", [])
                    ]
                    conversation_id = response.get("conversation_id")
            except Exception as exc:
                runtime_error = str(exc)

            retrieved_chunks = [
                compact_retrieval_item(item, rank)
                for rank, item in enumerate(retrieval_context, start=1)
            ]
            v2_case = v2_by_id.get(baseline["test_id"], {})
            evaluated.append(
                {
                    "test_id": baseline["test_id"],
                    "behaviour_category": baseline["behaviour_category"],
                    "question": baseline["question"],
                    "mentor_mode": baseline["mentor_mode"],
                    "course_id": COURSE_ID,
                    "route": route,
                    "retrieval_status": "CONTEXT"
                    if retrieved_chunks
                    else "NO_CONTEXT",
                    "retrieved_chunk_count": len(retrieved_chunks),
                    "retrieved_chunk_ids": [
                        item["chunk_id"] for item in retrieved_chunks
                    ],
                    "retrieved_chunk_titles": [
                        item["title"] for item in retrieved_chunks
                    ],
                    "retrieval_distances": [
                        item["distance"] for item in retrieved_chunks
                    ],
                    "retrieved_chunks": retrieved_chunks,
                    "prompt_v1_baseline_response": baseline.get(
                        "final_chatbot_response"
                    ),
                    "prompt_v2_response": v2_case.get("prompt_v2_chatbot_response"),
                    "prompt_v3_response": answer,
                    "sources": sources,
                    "conversation_id": conversation_id,
                    "api_runtime_error": runtime_error,
                    "llm_response_received": bool(answer) and runtime_error is None,
                }
            )

        return evaluated
    finally:
        retrieval_service.get_chroma_client = original_get_chroma_client
        shutil.rmtree(temp_dir, ignore_errors=True)


def summarize(evaluated: List[Dict[str, Any]]) -> Dict[str, Any]:
    return {
        "tests_executed": len(evaluated),
        "successful_llm_responses": sum(
            1 for item in evaluated if item["llm_response_received"]
        ),
        "api_runtime_errors": sum(1 for item in evaluated if item["api_runtime_error"]),
        "context_count": sum(
            1 for item in evaluated if item["retrieval_status"] == "CONTEXT"
        ),
        "no_context_count": sum(
            1 for item in evaluated if item["retrieval_status"] == "NO_CONTEXT"
        ),
        "target_test_ids": [item["test_id"] for item in evaluated],
    }


def write_workbook(evaluated: List[Dict[str, Any]], summary: Dict[str, Any]) -> None:
    workbook = openpyxl.Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "V3 Targeted Summary"
    summary_sheet.append(["Metric", "Value"])
    for key, value in summary.items():
        summary_sheet.append([key, ", ".join(value) if isinstance(value, list) else value])

    case_sheet = workbook.create_sheet("V3 Responses")
    case_sheet.append(
        [
            "Test ID",
            "Behaviour Category",
            "Question",
            "Mentor Mode",
            "Retrieval Status",
            "Retrieved Chunk Count",
            "Retrieved Chunk IDs",
            "Retrieved Chunk Titles",
            "Retrieval Distances",
            "Prompt V1 Response",
            "Prompt V2 Response",
            "Prompt V3 Response",
            "API Runtime Error",
        ]
    )
    for item in evaluated:
        case_sheet.append(
            [
                item["test_id"],
                item["behaviour_category"],
                item["question"],
                item["mentor_mode"],
                item["retrieval_status"],
                item["retrieved_chunk_count"],
                ", ".join(item["retrieved_chunk_ids"]) or "None",
                " | ".join(item["retrieved_chunk_titles"]) or "None",
                ", ".join(f"{distance:.6f}" for distance in item["retrieval_distances"])
                or "None",
                item["prompt_v1_baseline_response"] or "",
                item["prompt_v2_response"] or "",
                item["prompt_v3_response"] or "",
                item["api_runtime_error"] or "",
            ]
        )

    chunk_sheet = workbook.create_sheet("Retrieved Chunks")
    chunk_sheet.append(
        [
            "Test ID",
            "Rank",
            "Chunk ID",
            "Title",
            "Document Type",
            "Instructional Unit",
            "Distance",
        ]
    )
    for item in evaluated:
        if not item["retrieved_chunks"]:
            chunk_sheet.append([item["test_id"], "", "NO_CONTEXT", "", "", "", ""])
        for chunk in item["retrieved_chunks"]:
            chunk_sheet.append(
                [
                    item["test_id"],
                    chunk["rank"],
                    chunk["chunk_id"],
                    chunk["title"],
                    chunk["document_type"],
                    chunk["instructional_unit"],
                    round(chunk["distance"], 6),
                ]
            )

    workbook.save(OUTPUT_XLSX)


def write_json(evaluated: List[Dict[str, Any]], summary: Dict[str, Any]) -> None:
    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "stage": "Stage 3 - Prompt V3 Targeted Behaviour Test",
        "course_id": COURSE_ID,
        "default_top_k": 5,
        "max_chroma_distance": 1.4,
        "llm_called": True,
        "automated_scoring": False,
        "baseline_v1_source": str(BASELINE_V1_JSON),
        "baseline_v2_source": str(BASELINE_V2_JSON),
        "chroma_query_target": "temporary copy of existing ChromaDB contents",
        "summary": summary,
        "cases": evaluated,
    }
    OUTPUT_JSON.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def main() -> None:
    cases = load_cases()
    v2_by_id = load_v2_by_id()
    evaluated = run_targeted_v3(cases, v2_by_id)
    summary = summarize(evaluated)
    write_workbook(evaluated, summary)
    write_json(evaluated, summary)

    print(f"Wrote {OUTPUT_XLSX}")
    print(f"Wrote {OUTPUT_JSON}")
    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
