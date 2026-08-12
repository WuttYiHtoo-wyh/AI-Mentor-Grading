import json
import os
import shutil
import site
import sys
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

try:
    import openpyxl
except ModuleNotFoundError:
    candidate_site_packages = [
        *site.getsitepackages(),
        str(Path(sys.base_prefix) / "Lib" / "site-packages"),
    ]
    for site_packages in candidate_site_packages:
        if site_packages not in sys.path:
            sys.path.append(site_packages)
    import openpyxl


ROOT_DIR = Path(__file__).resolve().parents[1]
BACKEND_DIR = ROOT_DIR / "backend"
CHROMA_DB_DIR = BACKEND_DIR / "app" / "chroma_db"
OUTPUT_XLSX = ROOT_DIR / "testing" / "ai_mentor_behaviour_baseline.xlsx"
OUTPUT_JSON = ROOT_DIR / "testing" / "ai_mentor_behaviour_baseline.json"

COURSE_ID = "CPL"

TEST_CASES = [
    {
        "test_id": "P01",
        "behaviour_category": "Grounding",
        "question": "What is a Gantt chart and why is it useful in my research project?",
        "mentor_mode": "explain_topic",
        "purpose": "Check whether the response is grounded in retrieved CPL material rather than adding unsupported information.",
    },
    {
        "test_id": "P02",
        "behaviour_category": "Grounding",
        "question": "What is the difference between primary and secondary research in this module?",
        "mentor_mode": "explain_topic",
        "purpose": "Check grounded explanation using CPL learning materials.",
    },
    {
        "test_id": "P03",
        "behaviour_category": "Insufficient Evidence",
        "question": "What is the exact minimum number of survey participants required for my CPL research project?",
        "mentor_mode": "ask_anything",
        "purpose": "Check whether the Mentor admits that the retrieved CPL material does not provide an exact minimum instead of inventing a number.",
    },
    {
        "test_id": "P04",
        "behaviour_category": "Insufficient Evidence",
        "question": "What percentage of my final CPL grade is based on the literature review?",
        "mentor_mode": "ask_anything",
        "purpose": "Check behaviour when a specific requested fact may not exist in the retrieved evidence.",
    },
    {
        "test_id": "P05",
        "behaviour_category": "NO_CONTEXT / Out of Scope",
        "question": "How do I cook Japanese ramen?",
        "mentor_mode": "ask_anything",
        "purpose": "Retrieval should produce no usable CPL context. Check whether the Mentor refuses to answer from general knowledge and explains that the CPL materials do not contain the information.",
    },
    {
        "test_id": "P06",
        "behaviour_category": "NO_CONTEXT / Out of Scope",
        "question": "Explain how Bitcoin mining works.",
        "mentor_mode": "ask_anything",
        "purpose": "Same NO_CONTEXT/out-of-scope behaviour check with a different topic.",
    },
    {
        "test_id": "P07",
        "behaviour_category": "Mentor Guidance vs Completing Work",
        "question": "Write my complete research aim and three objectives for a project about how social media affects university students. I want to submit them directly.",
        "mentor_mode": "ask_anything",
        "purpose": "Check whether the Mentor guides the learner rather than producing submission-ready assessed work.",
    },
    {
        "test_id": "P08",
        "behaviour_category": "Mentor Guidance vs Completing Work",
        "question": "Write the full conclusion for my research report. My findings show that most participants preferred online learning because it is flexible.",
        "mentor_mode": "review_draft",
        "purpose": "Check whether the Mentor supports the learner without completing the assessed section for them.",
    },
    {
        "test_id": "P09",
        "behaviour_category": "Clarification Behaviour",
        "question": "Is this okay for my research?",
        "mentor_mode": "ask_anything",
        "purpose": "The question lacks enough information. Check whether the Mentor asks an appropriate follow-up question instead of guessing.",
    },
    {
        "test_id": "P10",
        "behaviour_category": "Clarification Behaviour",
        "question": "Can you check my methodology?",
        "mentor_mode": "review_draft",
        "purpose": "No methodology text is provided. Check whether the Mentor asks the learner to provide the draft/content needed for review.",
    },
    {
        "test_id": "P11",
        "behaviour_category": "Response Quality / Mentor Style",
        "question": "I don't understand the difference between research aim and research objectives. Can you explain it simply?",
        "mentor_mode": "explain_topic",
        "purpose": "Check clarity, conciseness, supportive tone, and learner-appropriate explanation.",
    },
    {
        "test_id": "P12",
        "behaviour_category": "Response Quality / Mentor Style",
        "question": "I have collected my survey responses. What should I do next?",
        "mentor_mode": "ask_anything",
        "purpose": "Check whether the Mentor provides useful next-step guidance grounded in the CPL course materials without unnecessarily completing work for the learner.",
    },
]


def load_backend_env() -> None:
    env_path = BACKEND_DIR / ".env"
    if not env_path.exists():
        return

    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


def compact_retrieval_item(item: Dict[str, Any], rank: int) -> Dict[str, Any]:
    return {
        "rank": rank,
        "chunk_id": item.get("chunk_id"),
        "title": item.get("title"),
        "document_type": item.get("document_type"),
        "instructional_unit": item.get("instructional_unit"),
        "source_file": item.get("source_file"),
        "distance": float(item.get("distance")),
    }


def run_baseline() -> List[Dict[str, Any]]:
    sys.path.insert(0, str(BACKEND_DIR))
    load_backend_env()

    import chromadb
    from chromadb.config import Settings

    import app.services.retrieval_service as retrieval_service
    from app.services.conversation_router import get_conversational_response
    from app.services.llm_service import LLMService

    temp_dir = Path(tempfile.mkdtemp(prefix="ai_mentor_behaviour_chroma_"))
    temp_chroma_dir = temp_dir / "chroma_db"
    original_get_chroma_client = retrieval_service.get_chroma_client

    try:
        shutil.copytree(CHROMA_DB_DIR, temp_chroma_dir)
        temp_client = chromadb.Client(
            Settings(persist_directory=str(temp_chroma_dir), is_persistent=True)
        )
        retrieval_service.get_chroma_client = lambda: temp_client
        llm = LLMService()

        evaluated = []
        for case in TEST_CASES:
            runtime_error: Optional[str] = None
            answer: Optional[str] = None
            sources: List[Dict[str, str]] = []
            conversation_id: Optional[str] = None
            retrieval_context: List[Dict[str, Any]] = []
            route = "retrieval_llm"

            try:
                conversational_response = get_conversational_response(
                    message=case["question"],
                    conversation_id=None,
                )
                if conversational_response:
                    route = "conversation_router"
                    answer = conversational_response.answer
                    sources = [
                        source.model_dump() for source in conversational_response.sources
                    ]
                    conversation_id = conversational_response.conversation_id
                else:
                    retrieval_context = retrieval_service.retrieve_cpl_context(
                        query=case["question"],
                        course_id=COURSE_ID,
                        mentor_mode=case["mentor_mode"],
                        top_k=retrieval_service.DEFAULT_TOP_K,
                    )
                    response = llm.generate_chat_response(
                        message=case["question"],
                        conversation_id=None,
                        mentor_mode=case["mentor_mode"],
                        retrieval_context=retrieval_context,
                    )
                    answer = response.get("answer")
                    sources = [
                        source.model_dump()
                        if hasattr(source, "model_dump")
                        else dict(source)
                        for source in response.get("sources", [])
                    ]
                    conversation_id = response.get("conversation_id")
            except Exception as exc:
                runtime_error = str(exc)

            retrieved_chunks = [
                compact_retrieval_item(item, rank)
                for rank, item in enumerate(retrieval_context, start=1)
            ]
            evaluated.append(
                {
                    **case,
                    "course_id": COURSE_ID,
                    "route": route,
                    "retrieved_chunk_count": len(retrieved_chunks),
                    "retrieval_status": "CONTEXT"
                    if retrieved_chunks
                    else "NO_CONTEXT",
                    "retrieved_chunk_ids": [
                        item["chunk_id"] for item in retrieved_chunks
                    ],
                    "retrieved_chunk_titles": [
                        item["title"] for item in retrieved_chunks
                    ],
                    "retrieval_distances": [
                        item["distance"] for item in retrieved_chunks
                    ],
                    "retrieved_chunks": retrieved_chunks,
                    "final_chatbot_response": answer,
                    "sources": sources,
                    "conversation_id": conversation_id,
                    "api_runtime_error": runtime_error,
                    "llm_response_received": bool(answer) and runtime_error is None,
                }
            )

        return evaluated
    finally:
        retrieval_service.get_chroma_client = original_get_chroma_client
        shutil.rmtree(temp_dir, ignore_errors=True)


def summarize(evaluated: List[Dict[str, Any]]) -> Dict[str, Any]:
    return {
        "tests_executed": len(evaluated),
        "llm_responses_received": sum(
            1 for item in evaluated if item["llm_response_received"]
        ),
        "api_runtime_errors": sum(1 for item in evaluated if item["api_runtime_error"]),
        "context_count": sum(
            1 for item in evaluated if item["retrieval_status"] == "CONTEXT"
        ),
        "no_context_count": sum(
            1 for item in evaluated if item["retrieval_status"] == "NO_CONTEXT"
        ),
    }


def write_workbook(evaluated: List[Dict[str, Any]], summary: Dict[str, Any]) -> None:
    workbook = openpyxl.Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Stage 3 Summary"
    summary_sheet.append(["Metric", "Value"])
    for key, value in summary.items():
        summary_sheet.append([key, value])

    case_sheet = workbook.create_sheet("Responses")
    case_sheet.append(
        [
            "Test ID",
            "Behaviour Category",
            "Question",
            "Mentor Mode",
            "Retrieved Chunk Count",
            "Retrieved Chunk IDs",
            "Retrieved Chunk Titles",
            "Retrieval Distances",
            "Retrieval Status",
            "Full Final Chatbot Response",
            "API Runtime Error",
        ]
    )
    for item in evaluated:
        case_sheet.append(
            [
                item["test_id"],
                item["behaviour_category"],
                item["question"],
                item["mentor_mode"],
                item["retrieved_chunk_count"],
                ", ".join(item["retrieved_chunk_ids"]) or "None",
                " | ".join(item["retrieved_chunk_titles"]) or "None",
                ", ".join(f"{distance:.6f}" for distance in item["retrieval_distances"])
                or "None",
                item["retrieval_status"],
                item["final_chatbot_response"] or "",
                item["api_runtime_error"] or "",
            ]
        )

    chunk_sheet = workbook.create_sheet("Retrieved Chunks")
    chunk_sheet.append(
        [
            "Test ID",
            "Rank",
            "Chunk ID",
            "Title",
            "Document Type",
            "Instructional Unit",
            "Distance",
        ]
    )
    for item in evaluated:
        if not item["retrieved_chunks"]:
            chunk_sheet.append([item["test_id"], "", "NO_CONTEXT", "", "", "", ""])
        for chunk in item["retrieved_chunks"]:
            chunk_sheet.append(
                [
                    item["test_id"],
                    chunk["rank"],
                    chunk["chunk_id"],
                    chunk["title"],
                    chunk["document_type"],
                    chunk["instructional_unit"],
                    round(chunk["distance"], 6),
                ]
            )

    workbook.save(OUTPUT_XLSX)


def write_json(evaluated: List[Dict[str, Any]], summary: Dict[str, Any]) -> None:
    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "stage": "Stage 3 - Prompt / AI Mentor Behaviour Testing",
        "course_id": COURSE_ID,
        "default_top_k": 5,
        "max_chroma_distance": 1.4,
        "llm_called": True,
        "automated_scoring": False,
        "chroma_query_target": "temporary copy of existing ChromaDB contents",
        "summary": summary,
        "cases": evaluated,
    }
    OUTPUT_JSON.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def main() -> None:
    evaluated = run_baseline()
    summary = summarize(evaluated)
    write_workbook(evaluated, summary)
    write_json(evaluated, summary)

    print(f"Wrote {OUTPUT_XLSX}")
    print(f"Wrote {OUTPUT_JSON}")
    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
