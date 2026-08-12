import json
import os
import shutil
import site
import sys
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

try:
    import openpyxl
except ModuleNotFoundError:
    candidate_site_packages = [
        *site.getsitepackages(),
        str(Path(sys.base_prefix) / "Lib" / "site-packages"),
    ]
    for site_packages in candidate_site_packages:
        if site_packages not in sys.path:
            sys.path.append(site_packages)
    import openpyxl


ROOT_DIR = Path(__file__).resolve().parents[1]
BACKEND_DIR = ROOT_DIR / "backend"
CHROMA_DB_DIR = BACKEND_DIR / "app" / "chroma_db"
OUTPUT_XLSX = ROOT_DIR / "testing" / "rag_mentor_mode_sanity_test.xlsx"
OUTPUT_JSON = ROOT_DIR / "testing" / "rag_mentor_mode_sanity_test.json"

COURSE_ID = "CPL"
TOP_K = 5
MAX_CHROMA_DISTANCE = 1.4

TEST_CASES = [
    {
        "test_id": "M01",
        "purpose": "explain_assignment mode filter",
        "question": "What do I need to submit for the CPL assignment?",
        "mentor_mode": "explain_assignment",
        "allowed_document_types": ["Assignment Brief"],
    },
    {
        "test_id": "M02",
        "purpose": "explain_rubric mode filter",
        "question": "What are the assessment criteria for the CPL assignment?",
        "mentor_mode": "explain_rubric",
        "allowed_document_types": ["Rubric"],
    },
    {
        "test_id": "M03",
        "purpose": "explain_topic mode filter",
        "question": "What is a Gantt chart?",
        "mentor_mode": "explain_topic",
        "allowed_document_types": ["Learning Material"],
    },
    {
        "test_id": "M04",
        "purpose": "review_draft allowed document types",
        "question": "Please review this draft section about my research question, survey method, and project schedule.",
        "mentor_mode": "review_draft",
        "allowed_document_types": ["Rubric", "Assignment Brief", "Learning Material"],
    },
    {
        "test_id": "M05",
        "purpose": "ask_anything course and threshold check",
        "question": "How should I plan my research project schedule?",
        "mentor_mode": "ask_anything",
        "allowed_document_types": None,
    },
    {
        "test_id": "Z01",
        "purpose": "zero-result and fallback-bypass probe",
        "question": "How do I cook Japanese ramen?",
        "mentor_mode": "explain_rubric",
        "allowed_document_types": ["Rubric"],
    },
]


def load_backend_env() -> None:
    env_path = BACKEND_DIR / ".env"
    if not env_path.exists():
        return

    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


def compact_item(item: Dict[str, Any], rank: int) -> Dict[str, Any]:
    metadata = item.get("metadata") or {}
    return {
        "rank": rank,
        "chunk_id": item.get("chunk_id"),
        "title": item.get("title"),
        "document_type": item.get("document_type"),
        "instructional_unit": item.get("instructional_unit"),
        "distance": float(item.get("distance")),
        "course_id": metadata.get("course_id"),
    }


def evaluate_case(case: Dict[str, Any], results: List[Dict[str, Any]]) -> Dict[str, Any]:
    chunks = [compact_item(item, rank) for rank, item in enumerate(results, start=1)]
    allowed_document_types: Optional[List[str]] = case["allowed_document_types"]
    course_failures = [item for item in chunks if item["course_id"] != COURSE_ID]
    threshold_failures = [
        item for item in chunks if item["distance"] > MAX_CHROMA_DISTANCE
    ]
    mode_failures = (
        []
        if allowed_document_types is None
        else [
            item
            for item in chunks
            if item["document_type"] not in allowed_document_types
        ]
    )
    returned_document_types = sorted(
        {item["document_type"] for item in chunks if item["document_type"]}
    )

    return {
        **case,
        "retrieved_count": len(chunks),
        "context_status": "CONTEXT" if chunks else "NO_CONTEXT",
        "retrieved_document_types": returned_document_types,
        "chunks": chunks,
        "course_filter_pass": not course_failures,
        "threshold_pass": not threshold_failures,
        "mode_filter_pass": not mode_failures,
        "overall_pass": not course_failures
        and not threshold_failures
        and not mode_failures,
        "course_failures": course_failures,
        "threshold_failures": threshold_failures,
        "mode_failures": mode_failures,
    }


def run_tests() -> List[Dict[str, Any]]:
    sys.path.insert(0, str(BACKEND_DIR))
    load_backend_env()

    import chromadb
    from chromadb.config import Settings

    import app.services.retrieval_service as retrieval_service

    temp_dir = Path(tempfile.mkdtemp(prefix="rag_mode_sanity_chroma_"))
    temp_chroma_dir = temp_dir / "chroma_db"
    original_get_chroma_client = retrieval_service.get_chroma_client

    try:
        shutil.copytree(CHROMA_DB_DIR, temp_chroma_dir)
        temp_client = chromadb.Client(
            Settings(persist_directory=str(temp_chroma_dir), is_persistent=True)
        )
        retrieval_service.get_chroma_client = lambda: temp_client

        evaluated = []
        for case in TEST_CASES:
            results = retrieval_service.retrieve_cpl_context(
                query=case["question"],
                course_id=COURSE_ID,
                mentor_mode=case["mentor_mode"],
                top_k=retrieval_service.DEFAULT_TOP_K,
            )
            evaluated.append(evaluate_case(case, results))

        return evaluated
    finally:
        retrieval_service.get_chroma_client = original_get_chroma_client
        shutil.rmtree(temp_dir, ignore_errors=True)


def summarize(evaluated: List[Dict[str, Any]]) -> Dict[str, Any]:
    requested_mode_cases = [case for case in evaluated if case["test_id"].startswith("M")]
    zero_probe = next(case for case in evaluated if case["test_id"] == "Z01")
    ask_anything = next(case for case in evaluated if case["test_id"] == "M05")

    return {
        "tested_configuration": {
            "default_top_k": TOP_K,
            "max_chroma_distance": MAX_CHROMA_DISTANCE,
            "course_id": COURSE_ID,
        },
        "mentor_mode_cases_passed": sum(
            1 for case in requested_mode_cases if case["overall_pass"]
        ),
        "mentor_mode_cases_total": len(requested_mode_cases),
        "failed_cases": [
            case["test_id"] for case in evaluated if not case["overall_pass"]
        ],
        "all_returned_results_course_id_cpl": all(
            case["course_filter_pass"] for case in evaluated
        ),
        "all_returned_results_distance_lte_1_4": all(
            case["threshold_pass"] for case in evaluated
        ),
        "no_mode_specific_filter_bypass_observed": all(
            case["mode_filter_pass"] for case in evaluated
        ),
        "zero_result_retrieval_possible": zero_probe["context_status"] == "NO_CONTEXT",
        "zero_result_probe": {
            "test_id": zero_probe["test_id"],
            "mentor_mode": zero_probe["mentor_mode"],
            "retrieved_count": zero_probe["retrieved_count"],
            "context_status": zero_probe["context_status"],
            "overall_pass": zero_probe["overall_pass"],
            "retrieved_document_types": zero_probe["retrieved_document_types"],
        },
        "ask_anything_document_types": ask_anything["retrieved_document_types"],
        "ask_anything_not_single_mode_restricted": len(
            ask_anything["retrieved_document_types"]
        )
        > 1,
        "case_results": [
            {
                "test_id": case["test_id"],
                "mentor_mode": case["mentor_mode"],
                "retrieved_count": case["retrieved_count"],
                "context_status": case["context_status"],
                "retrieved_document_types": case["retrieved_document_types"],
                "course_filter_pass": case["course_filter_pass"],
                "threshold_pass": case["threshold_pass"],
                "mode_filter_pass": case["mode_filter_pass"],
                "overall_pass": case["overall_pass"],
            }
            for case in evaluated
        ],
    }


def write_workbook(evaluated: List[Dict[str, Any]], summary: Dict[str, Any]) -> None:
    workbook = openpyxl.Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Mode Summary"
    summary_sheet.append(["Metric", "Value"])
    summary_sheet.append(["DEFAULT_TOP_K", TOP_K])
    summary_sheet.append(["MAX_CHROMA_DISTANCE", MAX_CHROMA_DISTANCE])
    summary_sheet.append(
        [
            "Mentor Mode Cases Passed",
            f"{summary['mentor_mode_cases_passed']} of {summary['mentor_mode_cases_total']}",
        ]
    )
    summary_sheet.append(
        ["Failed Cases", ", ".join(summary["failed_cases"]) or "None"]
    )
    summary_sheet.append(
        [
            "All Returned Results course_id=CPL",
            "Yes" if summary["all_returned_results_course_id_cpl"] else "No",
        ]
    )
    summary_sheet.append(
        [
            "All Returned Results distance<=1.4",
            "Yes" if summary["all_returned_results_distance_lte_1_4"] else "No",
        ]
    )
    summary_sheet.append(
        [
            "No Mode Filter Bypass Observed",
            "Yes" if summary["no_mode_specific_filter_bypass_observed"] else "No",
        ]
    )
    summary_sheet.append(
        [
            "Zero Result Retrieval Possible",
            "Yes" if summary["zero_result_retrieval_possible"] else "No",
        ]
    )
    summary_sheet.append(
        [
            "Ask Anything Document Types",
            ", ".join(summary["ask_anything_document_types"]) or "None",
        ]
    )

    case_sheet = workbook.create_sheet("Case Results")
    case_sheet.append(
        [
            "Test ID",
            "Purpose",
            "Question",
            "Mentor Mode",
            "Allowed Document Types",
            "Retrieved Count",
            "Context Status",
            "Retrieved Document Types",
            "Course Filter Pass",
            "Threshold Pass",
            "Mode Filter Pass",
            "Overall Pass",
        ]
    )
    for case in evaluated:
        case_sheet.append(
            [
                case["test_id"],
                case["purpose"],
                case["question"],
                case["mentor_mode"],
                ", ".join(case["allowed_document_types"])
                if case["allowed_document_types"]
                else "Any CPL type",
                case["retrieved_count"],
                case["context_status"],
                ", ".join(case["retrieved_document_types"]) or "None",
                "PASS" if case["course_filter_pass"] else "FAIL",
                "PASS" if case["threshold_pass"] else "FAIL",
                "PASS" if case["mode_filter_pass"] else "FAIL",
                "PASS" if case["overall_pass"] else "FAIL",
            ]
        )

    chunk_sheet = workbook.create_sheet("Retrieved Chunks")
    chunk_sheet.append(
        [
            "Test ID",
            "Question",
            "Mentor Mode",
            "Rank",
            "Chunk ID",
            "Title",
            "Document Type",
            "Instructional Unit",
            "Distance",
            "course_id",
            "Mode Filter Compliance",
        ]
    )
    for case in evaluated:
        if not case["chunks"]:
            chunk_sheet.append(
                [
                    case["test_id"],
                    case["question"],
                    case["mentor_mode"],
                    "",
                    "NO_CONTEXT",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "PASS",
                ]
            )
        for item in case["chunks"]:
            allowed = case["allowed_document_types"]
            mode_pass = allowed is None or item["document_type"] in allowed
            chunk_sheet.append(
                [
                    case["test_id"],
                    case["question"],
                    case["mentor_mode"],
                    item["rank"],
                    item["chunk_id"],
                    item["title"],
                    item["document_type"],
                    item["instructional_unit"],
                    round(item["distance"], 6),
                    item["course_id"],
                    "PASS" if mode_pass else "FAIL",
                ]
            )

    workbook.save(OUTPUT_XLSX)


def write_json(evaluated: List[Dict[str, Any]], summary: Dict[str, Any]) -> None:
    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "course_id": COURSE_ID,
        "top_k": TOP_K,
        "max_chroma_distance": MAX_CHROMA_DISTANCE,
        "llm_called": False,
        "embedding_model": "text-embedding-3-small",
        "chroma_collection": "cpl_documents",
        "chroma_query_target": "temporary copy of existing ChromaDB contents",
        "summary": summary,
        "cases": evaluated,
    }
    OUTPUT_JSON.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def main() -> None:
    evaluated = run_tests()
    summary = summarize(evaluated)
    write_workbook(evaluated, summary)
    write_json(evaluated, summary)

    print(f"Wrote {OUTPUT_XLSX}")
    print(f"Wrote {OUTPUT_JSON}")
    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
