import json
import os
import shutil
import site
import sys
import tempfile
from datetime import datetime
from pathlib import Path
from statistics import mean
from typing import Any, Dict, List

try:
    import openpyxl
except ModuleNotFoundError:
    candidate_site_packages = [
        *site.getsitepackages(),
        str(Path(sys.base_prefix) / "Lib" / "site-packages"),
    ]
    for site_packages in candidate_site_packages:
        if site_packages not in sys.path:
            sys.path.append(site_packages)
    import openpyxl


ROOT_DIR = Path(__file__).resolve().parents[1]
BACKEND_DIR = ROOT_DIR / "backend"
CHROMA_DB_DIR = BACKEND_DIR / "app" / "chroma_db"
OUTPUT_XLSX = ROOT_DIR / "testing" / "rag_scope_threshold_test.xlsx"
OUTPUT_JSON = ROOT_DIR / "testing" / "rag_scope_threshold_test.json"

COURSE_ID = "CPL"
MENTOR_MODE = "ask_anything"
COLLECTION_NAME = "cpl_documents"
TOP_K = 5
DISTANCE_THRESHOLD = 1.4

QUESTIONS = [
    {
        "test_id": "O01",
        "scope_group": "Clearly Out-of-Scope",
        "question": "How do I cook Japanese ramen?",
    },
    {
        "test_id": "O02",
        "scope_group": "Clearly Out-of-Scope",
        "question": "Who won the FIFA World Cup in 2022?",
    },
    {
        "test_id": "O03",
        "scope_group": "Clearly Out-of-Scope",
        "question": "Explain how Bitcoin mining works.",
    },
    {
        "test_id": "O04",
        "scope_group": "Clearly Out-of-Scope",
        "question": "How do I build a neural network for image recognition?",
    },
    {
        "test_id": "O05",
        "scope_group": "Clearly Out-of-Scope",
        "question": "What causes earthquakes?",
    },
    {
        "test_id": "B01",
        "scope_group": "Borderline / Potentially CPL-Relevant",
        "question": "Can I use Python to analyse my research data?",
    },
    {
        "test_id": "B02",
        "scope_group": "Borderline / Potentially CPL-Relevant",
        "question": "Can artificial intelligence be used as part of my research project?",
    },
    {
        "test_id": "B03",
        "scope_group": "Borderline / Potentially CPL-Relevant",
        "question": "How do I calculate percentages from my survey responses?",
    },
    {
        "test_id": "B04",
        "scope_group": "Borderline / Potentially CPL-Relevant",
        "question": "Can I use Excel to analyse and visualise my survey results?",
    },
    {
        "test_id": "B05",
        "scope_group": "Borderline / Potentially CPL-Relevant",
        "question": "What software should I use to create my research project schedule?",
    },
]


def load_backend_env() -> None:
    env_path = BACKEND_DIR / ".env"
    if not env_path.exists():
        return

    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


def rank_results(results: Dict[str, Any]) -> List[Dict[str, Any]]:
    documents = results.get("documents", [[]])[0]
    metadatas = results.get("metadatas", [[]])[0]
    distances = results.get("distances", [[]])[0]

    ranked_results = []
    for rank, (document, metadata, distance) in enumerate(
        zip(documents, metadatas, distances), start=1
    ):
        distance = float(distance)
        ranked_results.append(
            {
                "rank": rank,
                "chunk_id": metadata.get("chunk_id") or metadata.get("id") or "",
                "title": metadata.get("title") or metadata.get("heading") or "Unknown",
                "document_type": metadata.get("document_type") or "Unknown",
                "instructional_unit": metadata.get("instructional_unit"),
                "source_file": metadata.get("source_file") or metadata.get("source"),
                "distance": distance,
                "threshold_decision": "retained"
                if distance <= DISTANCE_THRESHOLD
                else "discarded",
            }
        )

    return ranked_results


def run_experiment() -> List[Dict[str, Any]]:
    sys.path.insert(0, str(BACKEND_DIR))
    load_backend_env()

    import chromadb
    from chromadb.config import Settings

    from app.services.embedding_service import EmbeddingService
    from app.services.retrieval_service import build_where_filter

    embedding_service = EmbeddingService()
    temp_dir = Path(tempfile.mkdtemp(prefix="rag_scope_chroma_"))
    temp_chroma_dir = temp_dir / "chroma_db"

    try:
        shutil.copytree(CHROMA_DB_DIR, temp_chroma_dir)
        client = chromadb.Client(
            Settings(persist_directory=str(temp_chroma_dir), is_persistent=True)
        )
        collection = client.get_collection(COLLECTION_NAME)
        where = build_where_filter(COURSE_ID, MENTOR_MODE)

        evaluated = []
        for case in QUESTIONS:
            query_embedding = embedding_service.create_embeddings([case["question"]])[0]
            raw_results = collection.query(
                query_embeddings=[query_embedding],
                n_results=TOP_K,
                where=where,
                include=["documents", "metadatas", "distances"],
            )
            retrieved_chunks = rank_results(raw_results)
            retained_chunks = [
                item
                for item in retrieved_chunks
                if item["threshold_decision"] == "retained"
            ]
            discarded_chunks = [
                item
                for item in retrieved_chunks
                if item["threshold_decision"] == "discarded"
            ]

            evaluated.append(
                {
                    **case,
                    "mentor_mode": MENTOR_MODE,
                    "top_k": TOP_K,
                    "distance_threshold": DISTANCE_THRESHOLD,
                    "original_count": len(retrieved_chunks),
                    "remaining_count": len(retained_chunks),
                    "context_status": "CONTEXT"
                    if retained_chunks
                    else "NO_CONTEXT",
                    "retrieved_chunks": retrieved_chunks,
                    "retained_chunks": retained_chunks,
                    "discarded_chunks": discarded_chunks,
                }
            )

        return evaluated
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)


def summarize(evaluated: List[Dict[str, Any]]) -> Dict[str, Any]:
    out_of_scope = [
        case for case in evaluated if case["scope_group"] == "Clearly Out-of-Scope"
    ]
    borderline = [
        case
        for case in evaluated
        if case["scope_group"] == "Borderline / Potentially CPL-Relevant"
    ]
    unexpected_cases = [
        case
        for case in evaluated
        if (
            case["scope_group"] == "Clearly Out-of-Scope"
            and case["context_status"] != "NO_CONTEXT"
        )
        or (
            case["scope_group"] == "Borderline / Potentially CPL-Relevant"
            and case["context_status"] != "CONTEXT"
        )
    ]

    return {
        "question_count": len(evaluated),
        "top_k": TOP_K,
        "distance_threshold": DISTANCE_THRESHOLD,
        "threshold_rule": "retain when distance <= 1.4",
        "average_chunks_before_filtering": mean(
            case["original_count"] for case in evaluated
        ),
        "average_chunks_after_filtering": mean(
            case["remaining_count"] for case in evaluated
        ),
        "out_of_scope_no_context_count": sum(
            1 for case in out_of_scope if case["context_status"] == "NO_CONTEXT"
        ),
        "out_of_scope_question_count": len(out_of_scope),
        "borderline_context_count": sum(
            1 for case in borderline if case["context_status"] == "CONTEXT"
        ),
        "borderline_question_count": len(borderline),
        "chunks_retained_by_question": [
            {
                "test_id": case["test_id"],
                "scope_group": case["scope_group"],
                "remaining_count": case["remaining_count"],
                "context_status": case["context_status"],
            }
            for case in evaluated
        ],
        "unexpected_cases": [
            {
                "test_id": case["test_id"],
                "scope_group": case["scope_group"],
                "context_status": case["context_status"],
                "remaining_count": case["remaining_count"],
                "distances": [
                    {
                        "rank": item["rank"],
                        "chunk_id": item["chunk_id"],
                        "distance": item["distance"],
                        "threshold_decision": item["threshold_decision"],
                    }
                    for item in case["retrieved_chunks"]
                ],
            }
            for case in unexpected_cases
        ],
    }


def write_workbook(evaluated: List[Dict[str, Any]], summary: Dict[str, Any]) -> None:
    workbook = openpyxl.Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Scope Summary"
    summary_sheet.append(["Metric", "Value"])
    summary_sheet.append(["Top K", TOP_K])
    summary_sheet.append(["Distance Threshold", DISTANCE_THRESHOLD])
    summary_sheet.append(["Threshold Rule", "retain when distance <= 1.4"])
    summary_sheet.append(
        [
            "Average Chunks Before Filtering",
            round(summary["average_chunks_before_filtering"], 2),
        ]
    )
    summary_sheet.append(
        [
            "Average Chunks After Filtering",
            round(summary["average_chunks_after_filtering"], 2),
        ]
    )
    summary_sheet.append(
        [
            "Clearly Out-of-Scope NO_CONTEXT",
            f"{summary['out_of_scope_no_context_count']} of {summary['out_of_scope_question_count']}",
        ]
    )
    summary_sheet.append(
        [
            "Borderline CONTEXT",
            f"{summary['borderline_context_count']} of {summary['borderline_question_count']}",
        ]
    )
    summary_sheet.append(
        [
            "Unexpected Cases",
            ", ".join(case["test_id"] for case in summary["unexpected_cases"])
            or "None",
        ]
    )

    case_sheet = workbook.create_sheet("Question Results")
    case_sheet.append(
        [
            "Test ID",
            "Scope Group",
            "Question",
            "Original Count",
            "Remaining Count",
            "Context Status",
            "Retained Chunk IDs",
            "Discarded Chunk IDs",
        ]
    )
    for case in evaluated:
        case_sheet.append(
            [
                case["test_id"],
                case["scope_group"],
                case["question"],
                case["original_count"],
                case["remaining_count"],
                case["context_status"],
                ", ".join(item["chunk_id"] for item in case["retained_chunks"])
                or "None",
                ", ".join(item["chunk_id"] for item in case["discarded_chunks"])
                or "None",
            ]
        )

    chunk_sheet = workbook.create_sheet("Retrieved Chunks")
    chunk_sheet.append(
        [
            "Test ID",
            "Scope Group",
            "Question",
            "Rank",
            "Chunk ID",
            "Title",
            "Document Type",
            "Instructional Unit",
            "Distance",
            "Threshold Decision",
            "Final Status",
        ]
    )
    for case in evaluated:
        for item in case["retrieved_chunks"]:
            chunk_sheet.append(
                [
                    case["test_id"],
                    case["scope_group"],
                    case["question"],
                    item["rank"],
                    item["chunk_id"],
                    item["title"],
                    item["document_type"],
                    item["instructional_unit"],
                    round(item["distance"], 6),
                    item["threshold_decision"],
                    case["context_status"],
                ]
            )

    workbook.save(OUTPUT_XLSX)


def write_json(evaluated: List[Dict[str, Any]], summary: Dict[str, Any]) -> None:
    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "course_id": COURSE_ID,
        "mentor_mode": MENTOR_MODE,
        "top_k": TOP_K,
        "distance_threshold": DISTANCE_THRESHOLD,
        "threshold_rule": "retain when distance <= 1.4",
        "llm_called": False,
        "relevance_classification": "none",
        "embedding_model": "text-embedding-3-small",
        "chroma_collection": COLLECTION_NAME,
        "chroma_source": str(CHROMA_DB_DIR),
        "chroma_query_target": "temporary copy of existing ChromaDB contents",
        "summary": summary,
        "cases": evaluated,
    }
    OUTPUT_JSON.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def main() -> None:
    evaluated = run_experiment()
    summary = summarize(evaluated)
    write_workbook(evaluated, summary)
    write_json(evaluated, summary)

    print(f"Wrote {OUTPUT_XLSX}")
    print(f"Wrote {OUTPUT_JSON}")
    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
