import json
import os
import shutil
import site
import sys
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List

try:
    import openpyxl
except ModuleNotFoundError:
    candidate_site_packages = [
        *site.getsitepackages(),
        str(Path(sys.base_prefix) / "Lib" / "site-packages"),
    ]
    for site_packages in candidate_site_packages:
        if site_packages not in sys.path:
            sys.path.append(site_packages)
    import openpyxl


ROOT_DIR = Path(__file__).resolve().parents[1]
BACKEND_DIR = ROOT_DIR / "backend"
CHROMA_DB_DIR = BACKEND_DIR / "app" / "chroma_db"
BASELINE_WORKBOOK = ROOT_DIR / "testing" / "AI_Mentor_Baseline_Chatbot_Testing.xlsx"
ORIGINAL_RETRIEVAL_JSON = ROOT_DIR / "testing" / "rag_combined_experiment_results.json"
OUTPUT_XLSX = ROOT_DIR / "testing" / "rag_retrieval_regression_results.xlsx"
OUTPUT_JSON = ROOT_DIR / "testing" / "rag_retrieval_regression_results.json"

COURSE_ID = "CPL"

MODE_MAP = {
    "Explain Assignment": "explain_assignment",
    "Explain Rubric": "explain_rubric",
    "Explain Topic": "explain_topic",
    "Review Draft": "review_draft",
    "Ask Anything": "ask_anything",
}


def load_backend_env() -> None:
    env_path = BACKEND_DIR / ".env"
    if not env_path.exists():
        return

    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


def mode_to_internal(label: str) -> str:
    if label not in MODE_MAP:
        raise ValueError(f"Unsupported baseline mentor mode: {label}")
    return MODE_MAP[label]


def load_baseline_cases() -> List[Dict[str, Any]]:
    workbook = openpyxl.load_workbook(BASELINE_WORKBOOK, data_only=True)
    worksheet = workbook["Baseline Tests"]
    cases = []

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        test_id = row[0]
        if not test_id:
            continue
        cases.append(
            {
                "test_id": test_id,
                "question_type": row[1],
                "cpl_area": row[2],
                "question": row[3],
                "expected_source_topic": row[4],
                "mentor_mode_label": row[5],
                "mentor_mode": mode_to_internal(row[5]),
                "baseline_note": row[19],
            }
        )

    if len(cases) != 15:
        raise RuntimeError(f"Expected 15 baseline cases, found {len(cases)}")

    return cases


def load_original_retrieval() -> Dict[str, Dict[str, Any]]:
    payload = json.loads(ORIGINAL_RETRIEVAL_JSON.read_text(encoding="utf-8"))
    cases = payload.get("cases", [])
    if len(cases) != 15:
        raise RuntimeError(f"Expected 15 original retrieval cases, found {len(cases)}")
    return {case["test_id"]: case for case in cases}


def compact_item(item: Dict[str, Any], rank: int = None) -> Dict[str, Any]:
    return {
        "rank": rank if rank is not None else item.get("rank"),
        "chunk_id": item.get("chunk_id"),
        "title": item.get("title"),
        "document_type": item.get("document_type"),
        "instructional_unit": item.get("instructional_unit"),
        "distance": float(item.get("distance")),
    }


def rank_updated_results(results: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    return [compact_item(item, rank=rank) for rank, item in enumerate(results, start=1)]


def run_updated_retrieval(cases: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    sys.path.insert(0, str(BACKEND_DIR))
    load_backend_env()

    import chromadb
    from chromadb.config import Settings

    import app.services.retrieval_service as retrieval_service

    temp_dir = Path(tempfile.mkdtemp(prefix="rag_regression_chroma_"))
    temp_chroma_dir = temp_dir / "chroma_db"
    original_get_chroma_client = retrieval_service.get_chroma_client

    try:
        shutil.copytree(CHROMA_DB_DIR, temp_chroma_dir)
        temp_client = chromadb.Client(
            Settings(persist_directory=str(temp_chroma_dir), is_persistent=True)
        )
        retrieval_service.get_chroma_client = lambda: temp_client

        evaluated = []
        for case in cases:
            results = retrieval_service.retrieve_cpl_context(
                query=case["question"],
                course_id=COURSE_ID,
                mentor_mode=case["mentor_mode"],
                top_k=retrieval_service.DEFAULT_TOP_K,
            )
            updated_chunks = rank_updated_results(results)
            evaluated.append(
                {
                    **case,
                    "updated_chunks": updated_chunks,
                    "updated_count": len(updated_chunks),
                    "context_status": "CONTEXT" if updated_chunks else "NO_CONTEXT",
                }
            )

        return evaluated
    finally:
        retrieval_service.get_chroma_client = original_get_chroma_client
        shutil.rmtree(temp_dir, ignore_errors=True)


def compare_cases(
    evaluated: List[Dict[str, Any]], original_by_id: Dict[str, Dict[str, Any]]
) -> List[Dict[str, Any]]:
    compared = []
    for case in evaluated:
        original_case = original_by_id[case["test_id"]]
        original_chunks = [
            compact_item(item) for item in original_case.get("retrieved_chunks", [])
        ]
        updated_chunks = case["updated_chunks"]

        original_ids = [item["chunk_id"] for item in original_chunks]
        updated_ids = [item["chunk_id"] for item in updated_chunks]
        removed_chunks = [
            item for item in original_chunks if item["chunk_id"] not in updated_ids
        ]
        added_chunks = [
            item for item in updated_chunks if item["chunk_id"] not in original_ids
        ]
        unchanged = original_ids == updated_ids
        changed_by_threshold_only = bool(removed_chunks) and not added_chunks and all(
            item["distance"] > 1.4 for item in removed_chunks
        )
        unexpected_change = (not unchanged) and not changed_by_threshold_only

        compared.append(
            {
                **case,
                "original_chunks": original_chunks,
                "original_count": len(original_chunks),
                "unchanged": unchanged,
                "removed_chunks": removed_chunks,
                "added_chunks": added_chunks,
                "changed_by_threshold_only": changed_by_threshold_only,
                "unexpected_change": unexpected_change,
                "potential_regression_note": (
                    "Previously available retrieval context was removed by the threshold. No human chunk-level relevance label is available in this test."
                    if removed_chunks and case["test_id"] != "T14"
                    else None
                ),
            }
        )

    return compared


def summarize(compared: List[Dict[str, Any]]) -> Dict[str, Any]:
    return {
        "total_test_cases": len(compared),
        "cases_with_unchanged_retrieval": [
            case["test_id"] for case in compared if case["unchanged"]
        ],
        "cases_changed_because_of_threshold": [
            case["test_id"] for case in compared if case["changed_by_threshold_only"]
        ],
        "cases_producing_no_context": [
            case["test_id"] for case in compared if case["context_status"] == "NO_CONTEXT"
        ],
        "unexpected_retrieval_changes": [
            case["test_id"] for case in compared if case["unexpected_change"]
        ],
        "potential_regressions_detected": [
            {
                "test_id": case["test_id"],
                "note": case["potential_regression_note"],
                "removed_chunks": case["removed_chunks"],
            }
            for case in compared
            if case["potential_regression_note"]
        ],
        "status_by_test": [
            {
                "test_id": case["test_id"],
                "original_count": case["original_count"],
                "updated_count": case["updated_count"],
                "context_status": case["context_status"],
                "removed_count": len(case["removed_chunks"]),
                "unchanged": case["unchanged"],
            }
            for case in compared
        ],
        "t01_status": next(case for case in compared if case["test_id"] == "T01"),
        "t04_status": next(case for case in compared if case["test_id"] == "T04"),
        "t14_status": next(case for case in compared if case["test_id"] == "T14"),
        "t15_status": next(case for case in compared if case["test_id"] == "T15"),
    }


def write_workbook(compared: List[Dict[str, Any]], summary: Dict[str, Any]) -> None:
    workbook = openpyxl.Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Regression Summary"
    summary_sheet.append(["Metric", "Value"])
    summary_sheet.append(["Total Test Cases", summary["total_test_cases"]])
    summary_sheet.append(
        [
            "Cases With Unchanged Retrieval",
            ", ".join(summary["cases_with_unchanged_retrieval"]) or "None",
        ]
    )
    summary_sheet.append(
        [
            "Cases Changed Because Of Threshold",
            ", ".join(summary["cases_changed_because_of_threshold"]) or "None",
        ]
    )
    summary_sheet.append(
        [
            "Cases Producing NO_CONTEXT",
            ", ".join(summary["cases_producing_no_context"]) or "None",
        ]
    )
    summary_sheet.append(
        [
            "Unexpected Retrieval Changes",
            ", ".join(summary["unexpected_retrieval_changes"]) or "None",
        ]
    )
    summary_sheet.append(
        [
            "Potential Regressions Detected",
            ", ".join(
                case["test_id"] for case in summary["potential_regressions_detected"]
            )
            or "None",
        ]
    )

    case_sheet = workbook.create_sheet("Case Comparison")
    case_sheet.append(
        [
            "Test ID",
            "Question",
            "Mentor Mode",
            "Original Count",
            "Updated Count",
            "Context Status",
            "Unchanged",
            "Changed By Threshold Only",
            "Unexpected Change",
            "Removed Chunk IDs",
            "Added Chunk IDs",
            "Potential Regression Note",
        ]
    )
    for case in compared:
        case_sheet.append(
            [
                case["test_id"],
                case["question"],
                case["mentor_mode"],
                case["original_count"],
                case["updated_count"],
                case["context_status"],
                "Yes" if case["unchanged"] else "No",
                "Yes" if case["changed_by_threshold_only"] else "No",
                "Yes" if case["unexpected_change"] else "No",
                ", ".join(item["chunk_id"] for item in case["removed_chunks"])
                or "None",
                ", ".join(item["chunk_id"] for item in case["added_chunks"]) or "None",
                case["potential_regression_note"] or "",
            ]
        )

    chunk_sheet = workbook.create_sheet("Updated Retrieval")
    chunk_sheet.append(
        [
            "Test ID",
            "Rank",
            "Chunk ID",
            "Title",
            "Document Type",
            "Instructional Unit",
            "Distance",
            "Context Status",
        ]
    )
    for case in compared:
        if not case["updated_chunks"]:
            chunk_sheet.append(
                [case["test_id"], "", "NO_CONTEXT", "", "", "", "", "NO_CONTEXT"]
            )
        for item in case["updated_chunks"]:
            chunk_sheet.append(
                [
                    case["test_id"],
                    item["rank"],
                    item["chunk_id"],
                    item["title"],
                    item["document_type"],
                    item["instructional_unit"],
                    round(item["distance"], 6),
                    case["context_status"],
                ]
            )

    removed_sheet = workbook.create_sheet("Removed Chunks")
    removed_sheet.append(
        [
            "Test ID",
            "Original Rank",
            "Chunk ID",
            "Title",
            "Document Type",
            "Instructional Unit",
            "Distance",
        ]
    )
    for case in compared:
        for item in case["removed_chunks"]:
            removed_sheet.append(
                [
                    case["test_id"],
                    item["rank"],
                    item["chunk_id"],
                    item["title"],
                    item["document_type"],
                    item["instructional_unit"],
                    round(item["distance"], 6),
                ]
            )

    workbook.save(OUTPUT_XLSX)


def write_json(compared: List[Dict[str, Any]], summary: Dict[str, Any]) -> None:
    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "course_id": COURSE_ID,
        "top_k": 5,
        "max_chroma_distance": 1.4,
        "llm_called": False,
        "relevance_classification": "none",
        "original_retrieval_source": str(ORIGINAL_RETRIEVAL_JSON),
        "chroma_query_target": "temporary copy of existing ChromaDB contents",
        "summary": summary,
        "cases": compared,
    }
    OUTPUT_JSON.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def main() -> None:
    baseline_cases = load_baseline_cases()
    original_by_id = load_original_retrieval()
    evaluated = run_updated_retrieval(baseline_cases)
    compared = compare_cases(evaluated, original_by_id)
    summary = summarize(compared)

    write_workbook(compared, summary)
    write_json(compared, summary)

    print(f"Wrote {OUTPUT_XLSX}")
    print(f"Wrote {OUTPUT_JSON}")
    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
