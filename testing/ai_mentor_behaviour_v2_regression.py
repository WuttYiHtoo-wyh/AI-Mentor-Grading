import json
import os
import shutil
import site
import sys
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

try:
    import openpyxl
except ModuleNotFoundError:
    candidate_site_packages = [
        *site.getsitepackages(),
        str(Path(sys.base_prefix) / "Lib" / "site-packages"),
    ]
    for site_packages in candidate_site_packages:
        if site_packages not in sys.path:
            sys.path.append(site_packages)
    import openpyxl


ROOT_DIR = Path(__file__).resolve().parents[1]
BACKEND_DIR = ROOT_DIR / "backend"
CHROMA_DB_DIR = BACKEND_DIR / "app" / "chroma_db"
BASELINE_JSON = ROOT_DIR / "testing" / "ai_mentor_behaviour_baseline.json"
OUTPUT_XLSX = ROOT_DIR / "testing" / "ai_mentor_behaviour_v2_regression.xlsx"
OUTPUT_JSON = ROOT_DIR / "testing" / "ai_mentor_behaviour_v2_regression.json"

COURSE_ID = "CPL"


def load_backend_env() -> None:
    env_path = BACKEND_DIR / ".env"
    if not env_path.exists():
        return

    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


def load_baseline_cases() -> List[Dict[str, Any]]:
    payload = json.loads(BASELINE_JSON.read_text(encoding="utf-8"))
    cases = payload.get("cases", [])
    expected_ids = [f"P{index:02d}" for index in range(1, 13)]
    actual_ids = [case.get("test_id") for case in cases]
    if actual_ids != expected_ids:
        raise RuntimeError(f"Baseline cases are not exactly P01-P12: {actual_ids}")
    return cases


def compact_retrieval_item(item: Dict[str, Any], rank: int) -> Dict[str, Any]:
    return {
        "rank": rank,
        "chunk_id": item.get("chunk_id"),
        "title": item.get("title"),
        "document_type": item.get("document_type"),
        "instructional_unit": item.get("instructional_unit"),
        "source_file": item.get("source_file"),
        "distance": float(item.get("distance")),
    }


def run_v2_regression(baseline_cases: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    sys.path.insert(0, str(BACKEND_DIR))
    load_backend_env()

    import chromadb
    from chromadb.config import Settings

    import app.services.retrieval_service as retrieval_service
    from app.services.conversation_router import get_conversational_response
    from app.services.llm_service import LLMService

    temp_dir = Path(tempfile.mkdtemp(prefix="ai_mentor_v2_chroma_"))
    temp_chroma_dir = temp_dir / "chroma_db"
    original_get_chroma_client = retrieval_service.get_chroma_client

    try:
        shutil.copytree(CHROMA_DB_DIR, temp_chroma_dir)
        temp_client = chromadb.Client(
            Settings(persist_directory=str(temp_chroma_dir), is_persistent=True)
        )
        retrieval_service.get_chroma_client = lambda: temp_client
        llm = LLMService()

        evaluated = []
        for baseline in baseline_cases:
            runtime_error: Optional[str] = None
            answer: Optional[str] = None
            sources: List[Dict[str, str]] = []
            conversation_id: Optional[str] = None
            retrieval_context: List[Dict[str, Any]] = []
            route = "retrieval_llm"

            try:
                conversational_response = get_conversational_response(
                    message=baseline["question"],
                    conversation_id=None,
                )
                if conversational_response:
                    route = "conversation_router"
                    answer = conversational_response.answer
                    sources = [
                        source.model_dump() for source in conversational_response.sources
                    ]
                    conversation_id = conversational_response.conversation_id
                else:
                    retrieval_context = retrieval_service.retrieve_cpl_context(
                        query=baseline["question"],
                        course_id=COURSE_ID,
                        mentor_mode=baseline["mentor_mode"],
                        top_k=retrieval_service.DEFAULT_TOP_K,
                    )
                    response = llm.generate_chat_response(
                        message=baseline["question"],
                        conversation_id=None,
                        mentor_mode=baseline["mentor_mode"],
                        retrieval_context=retrieval_context,
                    )
                    answer = response.get("answer")
                    sources = [
                        source.model_dump()
                        if hasattr(source, "model_dump")
                        else dict(source)
                        for source in response.get("sources", [])
                    ]
                    conversation_id = response.get("conversation_id")
            except Exception as exc:
                runtime_error = str(exc)

            retrieved_chunks = [
                compact_retrieval_item(item, rank)
                for rank, item in enumerate(retrieval_context, start=1)
            ]
            evaluated.append(
                {
                    "test_id": baseline["test_id"],
                    "behaviour_category": baseline["behaviour_category"],
                    "question": baseline["question"],
                    "mentor_mode": baseline["mentor_mode"],
                    "course_id": COURSE_ID,
                    "route": route,
                    "retrieved_chunk_count": len(retrieved_chunks),
                    "retrieval_status": "CONTEXT"
                    if retrieved_chunks
                    else "NO_CONTEXT",
                    "retrieved_chunk_ids": [
                        item["chunk_id"] for item in retrieved_chunks
                    ],
                    "retrieved_chunk_titles": [
                        item["title"] for item in retrieved_chunks
                    ],
                    "retrieval_distances": [
                        item["distance"] for item in retrieved_chunks
                    ],
                    "retrieved_chunks": retrieved_chunks,
                    "prompt_v1_baseline_response": baseline.get(
                        "final_chatbot_response"
                    ),
                    "prompt_v2_chatbot_response": answer,
                    "sources": sources,
                    "conversation_id": conversation_id,
                    "api_runtime_error": runtime_error,
                    "llm_response_received": bool(answer) and runtime_error is None,
                }
            )

        return evaluated
    finally:
        retrieval_service.get_chroma_client = original_get_chroma_client
        shutil.rmtree(temp_dir, ignore_errors=True)


def summarize(
    baseline_cases: List[Dict[str, Any]], evaluated: List[Dict[str, Any]]
) -> Dict[str, Any]:
    all_questions_exactly_matched = all(
        baseline["test_id"] == result["test_id"]
        and baseline["question"] == result["question"]
        and baseline["mentor_mode"] == result["mentor_mode"]
        for baseline, result in zip(baseline_cases, evaluated)
    )
    return {
        "tests_executed": len(evaluated),
        "successful_llm_responses": sum(
            1 for item in evaluated if item["llm_response_received"]
        ),
        "api_runtime_errors": sum(1 for item in evaluated if item["api_runtime_error"]),
        "context_count": sum(
            1 for item in evaluated if item["retrieval_status"] == "CONTEXT"
        ),
        "no_context_count": sum(
            1 for item in evaluated if item["retrieval_status"] == "NO_CONTEXT"
        ),
        "all_p01_p12_questions_exactly_matched_baseline": all_questions_exactly_matched,
    }


def write_workbook(evaluated: List[Dict[str, Any]], summary: Dict[str, Any]) -> None:
    workbook = openpyxl.Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "V2 Summary"
    summary_sheet.append(["Metric", "Value"])
    for key, value in summary.items():
        summary_sheet.append([key, value])

    case_sheet = workbook.create_sheet("V2 Responses")
    case_sheet.append(
        [
            "Test ID",
            "Behaviour Category",
            "Question",
            "Mentor Mode",
            "Retrieval Status",
            "Retrieved Chunk Count",
            "Retrieved Chunk IDs",
            "Retrieved Chunk Titles",
            "Retrieval Distances",
            "Prompt V1 Baseline Response",
            "Prompt V2 Chatbot Response",
            "API Runtime Error",
        ]
    )
    for item in evaluated:
        case_sheet.append(
            [
                item["test_id"],
                item["behaviour_category"],
                item["question"],
                item["mentor_mode"],
                item["retrieval_status"],
                item["retrieved_chunk_count"],
                ", ".join(item["retrieved_chunk_ids"]) or "None",
                " | ".join(item["retrieved_chunk_titles"]) or "None",
                ", ".join(f"{distance:.6f}" for distance in item["retrieval_distances"])
                or "None",
                item["prompt_v1_baseline_response"] or "",
                item["prompt_v2_chatbot_response"] or "",
                item["api_runtime_error"] or "",
            ]
        )

    chunk_sheet = workbook.create_sheet("Retrieved Chunks")
    chunk_sheet.append(
        [
            "Test ID",
            "Rank",
            "Chunk ID",
            "Title",
            "Document Type",
            "Instructional Unit",
            "Distance",
        ]
    )
    for item in evaluated:
        if not item["retrieved_chunks"]:
            chunk_sheet.append([item["test_id"], "", "NO_CONTEXT", "", "", "", ""])
        for chunk in item["retrieved_chunks"]:
            chunk_sheet.append(
                [
                    item["test_id"],
                    chunk["rank"],
                    chunk["chunk_id"],
                    chunk["title"],
                    chunk["document_type"],
                    chunk["instructional_unit"],
                    round(chunk["distance"], 6),
                ]
            )

    workbook.save(OUTPUT_XLSX)


def write_json(evaluated: List[Dict[str, Any]], summary: Dict[str, Any]) -> None:
    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "stage": "Stage 3 - Prompt V2 Behaviour Regression",
        "course_id": COURSE_ID,
        "default_top_k": 5,
        "max_chroma_distance": 1.4,
        "llm_called": True,
        "automated_scoring": False,
        "baseline_source": str(BASELINE_JSON),
        "chroma_query_target": "temporary copy of existing ChromaDB contents",
        "summary": summary,
        "cases": evaluated,
    }
    OUTPUT_JSON.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def main() -> None:
    baseline_cases = load_baseline_cases()
    evaluated = run_v2_regression(baseline_cases)
    summary = summarize(baseline_cases, evaluated)
    write_workbook(evaluated, summary)
    write_json(evaluated, summary)

    print(f"Wrote {OUTPUT_XLSX}")
    print(f"Wrote {OUTPUT_JSON}")
    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
