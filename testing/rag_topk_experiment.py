import json
import os
import re
import site
import sys
from datetime import datetime
from pathlib import Path
from statistics import mean
from typing import Any, Dict, List, Optional

try:
    import openpyxl
except ModuleNotFoundError:
    candidate_site_packages = [
        *site.getsitepackages(),
        str(Path(sys.base_prefix) / "Lib" / "site-packages"),
    ]
    for site_packages in candidate_site_packages:
        if site_packages not in sys.path:
            sys.path.append(site_packages)
    import openpyxl


ROOT_DIR = Path(__file__).resolve().parents[1]
BACKEND_DIR = ROOT_DIR / "backend"
BASELINE_WORKBOOK = ROOT_DIR / "testing" / "AI_Mentor_Baseline_Chatbot_Testing.xlsx"
OUTPUT_XLSX = ROOT_DIR / "testing" / "rag_topk_experiment_results.xlsx"
OUTPUT_JSON = ROOT_DIR / "testing" / "rag_topk_experiment_results.json"

COURSE_ID = "CPL"
TOP_K_VALUES = [3, 5]

MODE_MAP = {
    "Explain Assignment": "explain_assignment",
    "Explain Rubric": "explain_rubric",
    "Explain Topic": "explain_topic",
    "Review Draft": "review_draft",
    "Ask Anything": "ask_anything",
}

# Query-specific indicators are used only for offline retrieval analysis.
# They do not affect retrieval, ranking, prompts, embeddings, or app config.
POTENTIAL_RELEVANCE_INDICATORS = {
    "T01": ["gantt"],
    "T02": ["gantt", "work breakdown", "wbs", "dependency", "dependencies", "schedule"],
    "T03": ["research question", "too broad", "scope", "specific", "literature review"],
    "T04": ["sample", "sampling", "participant", "survey", "methodology", "limitation", "response"],
    "T05": ["research topic", "topic selection", "scope", "feasible", "interest"],
    "T06": ["literature review", "source", "gap", "background", "existing research"],
    "T07": ["qualitative", "quantitative", "methodology", "experience", "detail"],
    "T08": ["primary research", "secondary research", "survey", "websites", "reports"],
    "T09": ["data analysis", "prepare data", "clean", "survey data", "findings"],
    "T10": ["finding", "recommendation"],
    "T11": ["reflection", "future research", "limitation", "evaluate"],
    "T12": ["alignment", "research question", "survey question", "finding", "recommendation"],
    "T13": ["project planning", "planning", "wbs", "gantt", "schedule", "milestone", "risk"],
    "T14": [],
    "T15": ["survey", "participant", "sample", "respondent", "questionnaire", "data collection"],
}


def load_backend_env() -> None:
    env_path = BACKEND_DIR / ".env"
    if not env_path.exists():
        return

    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


def normalize_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip().lower()


def mode_to_internal(label: str) -> str:
    if label not in MODE_MAP:
        raise ValueError(f"Unsupported baseline mentor mode: {label}")
    return MODE_MAP[label]


def load_baseline_cases() -> List[Dict[str, Any]]:
    workbook = openpyxl.load_workbook(BASELINE_WORKBOOK, data_only=True)
    worksheet = workbook["Baseline Tests"]
    cases = []

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        test_id = row[0]
        if not test_id:
            continue
        cases.append(
            {
                "test_id": test_id,
                "question_type": row[1],
                "cpl_area": row[2],
                "question": row[3],
                "expected_source_topic": row[4],
                "mentor_mode_label": row[5],
                "mentor_mode": mode_to_internal(row[5]),
                "baseline_note": row[19],
            }
        )

    if len(cases) != 15:
        raise RuntimeError(f"Expected 15 baseline cases, found {len(cases)}")

    return cases


def is_potentially_relevant(test_id: str, item: Dict[str, Any]) -> bool:
    indicators = POTENTIAL_RELEVANCE_INDICATORS.get(test_id, [])
    if not indicators:
        return False

    metadata = item.get("metadata") or {}
    haystack = normalize_text(
        " ".join(
            str(value or "")
            for value in [
                item.get("title"),
                item.get("instructional_unit"),
                item.get("document_type"),
                item.get("source_file"),
                item.get("content"),
                metadata.get("keywords"),
                metadata.get("question"),
                metadata.get("topic"),
                metadata.get("section"),
            ]
        )
    )
    return any(indicator in haystack for indicator in indicators)


def rank_results(test_id: str, results: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    ranked_results = []
    for rank, item in enumerate(results, start=1):
        distance = float(item.get("distance"))
        ranked_results.append(
            {
                "rank": rank,
                "chunk_id": item.get("chunk_id"),
                "title": item.get("title"),
                "instructional_unit": item.get("instructional_unit"),
                "document_type": item.get("document_type"),
                "source_file": item.get("source_file"),
                "distance": distance,
                "potentially_relevant": is_potentially_relevant(test_id, item),
            }
        )
    return ranked_results


def first_relevant(results: List[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    return next((item for item in results if item["potentially_relevant"]), None)


def retrieve_cases(cases: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    sys.path.insert(0, str(BACKEND_DIR))
    load_backend_env()

    from app.services.retrieval_service import retrieve_cpl_context

    evaluated = []
    for case in cases:
        top3 = rank_results(
            case["test_id"],
            retrieve_cpl_context(
                query=case["question"],
                course_id=COURSE_ID,
                mentor_mode=case["mentor_mode"],
                top_k=3,
            ),
        )
        top5 = rank_results(
            case["test_id"],
            retrieve_cpl_context(
                query=case["question"],
                course_id=COURSE_ID,
                mentor_mode=case["mentor_mode"],
                top_k=5,
            ),
        )

        top3_first = first_relevant(top3)
        top5_first = first_relevant(top5)
        ranks_4_5 = [item for item in top5 if item["rank"] in (4, 5)]
        useful_4_5 = [item for item in ranks_4_5 if item["potentially_relevant"]]
        noisy_4_5 = [item for item in ranks_4_5 if not item["potentially_relevant"]]

        evaluated.append(
            {
                **case,
                "top3_results": top3,
                "top5_results": top5,
                "top3_prefix_matches_top5": [item["chunk_id"] for item in top3]
                == [item["chunk_id"] for item in top5[:3]],
                "most_relevant_same_rank": (
                    top3_first is not None
                    and top5_first is not None
                    and top3_first["chunk_id"] == top5_first["chunk_id"]
                    and top3_first["rank"] == top5_first["rank"]
                ),
                "top3_first_relevant": top3_first,
                "top5_first_relevant": top5_first,
                "rank4_5_useful_supporting": useful_4_5,
                "rank4_5_noisy": noisy_4_5,
                "top3_would_lose_useful_evidence": bool(useful_4_5),
                "top5_only_adds_noise": bool(ranks_4_5) and not useful_4_5,
                "out_of_scope_behavior": (
                    "T14 out-of-scope query still retrieved CPL chunks; all retrieved chunks are treated as noise for this experiment."
                    if case["test_id"] == "T14"
                    else None
                ),
            }
        )

    return evaluated


def summarize(evaluated: List[Dict[str, Any]]) -> Dict[str, Any]:
    top3_relevant_counts = [
        sum(1 for item in case["top3_results"] if item["potentially_relevant"])
        for case in evaluated
    ]
    top5_relevant_counts = [
        sum(1 for item in case["top5_results"] if item["potentially_relevant"])
        for case in evaluated
    ]
    rank4_5_relevant_counts = [
        len(case["rank4_5_useful_supporting"]) for case in evaluated
    ]
    rank4_5_noise_counts = [len(case["rank4_5_noisy"]) for case in evaluated]

    return {
        "question_count": len(evaluated),
        "top3": {
            "useful_evidence_retained_chunks": sum(top3_relevant_counts),
            "average_useful_chunks_per_query": mean(top3_relevant_counts),
            "irrelevant_chunks_retrieved": sum(
                len(case["top3_results"])
                - sum(1 for item in case["top3_results"] if item["potentially_relevant"])
                for case in evaluated
            ),
            "potential_information_loss_cases": [
                case["test_id"]
                for case in evaluated
                if case["top3_would_lose_useful_evidence"]
            ],
        },
        "top5": {
            "useful_evidence_chunks": sum(top5_relevant_counts),
            "average_useful_chunks_per_query": mean(top5_relevant_counts),
            "additional_useful_rank4_5_chunks": sum(rank4_5_relevant_counts),
            "additional_irrelevant_rank4_5_chunks": sum(rank4_5_noise_counts),
            "materially_improved_coverage_cases": [
                case["test_id"]
                for case in evaluated
                if case["rank4_5_useful_supporting"]
            ],
            "only_added_noise_cases": [
                case["test_id"] for case in evaluated if case["top5_only_adds_noise"]
            ],
        },
        "rank_stability": {
            "top3_prefix_matched_top5_cases": [
                case["test_id"] for case in evaluated if case["top3_prefix_matches_top5"]
            ],
            "most_relevant_same_rank_cases": [
                case["test_id"]
                for case in evaluated
                if case["most_relevant_same_rank"]
            ],
            "no_identified_relevant_chunk_cases": [
                case["test_id"]
                for case in evaluated
                if case["top3_first_relevant"] is None and case["top5_first_relevant"] is None
            ],
        },
        "out_of_scope": {
            "t14": next(case for case in evaluated if case["test_id"] == "T14")[
                "out_of_scope_behavior"
            ],
        },
    }


def write_workbook(evaluated: List[Dict[str, Any]], summary: Dict[str, Any]) -> None:
    workbook = openpyxl.Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "TopK Summary"
    summary_sheet.append(["Metric", "Top 3", "Top 5"])
    summary_sheet.append(
        [
            "Useful evidence chunks",
            summary["top3"]["useful_evidence_retained_chunks"],
            summary["top5"]["useful_evidence_chunks"],
        ]
    )
    summary_sheet.append(
        [
            "Average useful chunks per query",
            round(summary["top3"]["average_useful_chunks_per_query"], 2),
            round(summary["top5"]["average_useful_chunks_per_query"], 2),
        ]
    )
    summary_sheet.append(
        [
            "Irrelevant chunks",
            summary["top3"]["irrelevant_chunks_retrieved"],
            summary["top5"]["additional_irrelevant_rank4_5_chunks"],
        ]
    )
    summary_sheet.append(
        [
            "Cases with potential Top 3 information loss",
            ", ".join(summary["top3"]["potential_information_loss_cases"]) or "None",
            "",
        ]
    )
    summary_sheet.append(
        [
            "Cases where Top 5 materially improves coverage",
            "",
            ", ".join(summary["top5"]["materially_improved_coverage_cases"]) or "None",
        ]
    )
    summary_sheet.append(
        [
            "Cases where Top 5 only adds noise",
            "",
            ", ".join(summary["top5"]["only_added_noise_cases"]) or "None",
        ]
    )
    summary_sheet.append(["Out-of-scope T14 behavior", "", summary["out_of_scope"]["t14"]])

    comparison_sheet = workbook.create_sheet("Case Comparison")
    comparison_sheet.append(
        [
            "Test ID",
            "Question",
            "Mentor Mode",
            "Top 3 Prefix Matches Top 5",
            "Most Relevant Same Rank",
            "Top 3 Would Lose Useful Evidence",
            "Top 5 Only Adds Noise",
            "Useful Ranks 4-5",
            "Noisy Ranks 4-5",
            "Out-of-Scope Behavior",
        ]
    )
    for case in evaluated:
        comparison_sheet.append(
            [
                case["test_id"],
                case["question"],
                case["mentor_mode"],
                "Yes" if case["top3_prefix_matches_top5"] else "No",
                "Yes" if case["most_relevant_same_rank"] else "No",
                "Yes" if case["top3_would_lose_useful_evidence"] else "No",
                "Yes" if case["top5_only_adds_noise"] else "No",
                ", ".join(
                    f"r{item['rank']} {item['chunk_id']}"
                    for item in case["rank4_5_useful_supporting"]
                )
                or "None",
                ", ".join(
                    f"r{item['rank']} {item['chunk_id']}" for item in case["rank4_5_noisy"]
                )
                or "None",
                case["out_of_scope_behavior"] or "",
            ]
        )

    detail_sheet = workbook.create_sheet("Retrieved Chunks")
    detail_sheet.append(
        [
            "Test ID",
            "Question",
            "Mentor Mode",
            "Top K",
            "Rank",
            "Chunk ID",
            "Title",
            "Instructional Unit",
            "Document Type",
            "Distance",
            "Potentially Relevant",
        ]
    )
    for case in evaluated:
        for top_k, key in [(3, "top3_results"), (5, "top5_results")]:
            for item in case[key]:
                detail_sheet.append(
                    [
                        case["test_id"],
                        case["question"],
                        case["mentor_mode"],
                        top_k,
                        item["rank"],
                        item["chunk_id"],
                        item["title"],
                        item["instructional_unit"],
                        item["document_type"],
                        round(item["distance"], 6),
                        "Yes" if item["potentially_relevant"] else "No",
                    ]
                )

    workbook.save(OUTPUT_XLSX)


def write_json(evaluated: List[Dict[str, Any]], summary: Dict[str, Any]) -> None:
    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "course_id": COURSE_ID,
        "top_k_values": TOP_K_VALUES,
        "threshold_applied": False,
        "llm_called": False,
        "embedding_model": "text-embedding-3-small",
        "chroma_collection": "cpl_documents",
        "summary": summary,
        "cases": evaluated,
    }
    OUTPUT_JSON.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def main() -> None:
    cases = load_baseline_cases()
    evaluated = retrieve_cases(cases)
    summary = summarize(evaluated)

    write_workbook(evaluated, summary)
    write_json(evaluated, summary)

    print(f"Wrote {OUTPUT_XLSX}")
    print(f"Wrote {OUTPUT_JSON}")
    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
