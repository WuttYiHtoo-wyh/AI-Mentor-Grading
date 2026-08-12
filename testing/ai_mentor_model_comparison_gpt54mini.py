import json
import os
import shutil
import site
import sys
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

try:
    import openpyxl
except ModuleNotFoundError:
    candidate_site_packages = [
        *site.getsitepackages(),
        str(Path(sys.base_prefix) / "Lib" / "site-packages"),
    ]
    for site_packages in candidate_site_packages:
        if site_packages not in sys.path:
            sys.path.append(site_packages)
    import openpyxl


ROOT_DIR = Path(__file__).resolve().parents[1]
BACKEND_DIR = ROOT_DIR / "backend"
CHROMA_DB_DIR = BACKEND_DIR / "app" / "chroma_db"
BASELINE_V1_JSON = ROOT_DIR / "testing" / "ai_mentor_behaviour_baseline.json"
BASELINE_V3_JSON = ROOT_DIR / "testing" / "ai_mentor_behaviour_v3_targeted.json"
OUTPUT_XLSX = ROOT_DIR / "testing" / "ai_mentor_model_comparison_gpt54mini.xlsx"
OUTPUT_JSON = ROOT_DIR / "testing" / "ai_mentor_model_comparison_gpt54mini.json"

COURSE_ID = "CPL"
MODEL = "gpt-5.4-mini"
MAX_COMPLETION_TOKENS = 500
TEMPERATURE = 0.2
TARGET_IDS = ["P05", "P07", "P08", "P09", "P10"]


def load_backend_env() -> None:
    env_path = BACKEND_DIR / ".env"
    if not env_path.exists():
        return

    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


def load_cases() -> List[Dict[str, Any]]:
    payload = json.loads(BASELINE_V1_JSON.read_text(encoding="utf-8"))
    cases_by_id = {case["test_id"]: case for case in payload.get("cases", [])}
    missing = [test_id for test_id in TARGET_IDS if test_id not in cases_by_id]
    if missing:
        raise RuntimeError(f"Missing target cases in V1 baseline: {missing}")
    return [cases_by_id[test_id] for test_id in TARGET_IDS]


def load_v3_by_id() -> Dict[str, Dict[str, Any]]:
    if not BASELINE_V3_JSON.exists():
        return {}
    payload = json.loads(BASELINE_V3_JSON.read_text(encoding="utf-8"))
    return {case["test_id"]: case for case in payload.get("cases", [])}


def compact_retrieval_item(item: Dict[str, Any], rank: int) -> Dict[str, Any]:
    return {
        "rank": rank,
        "chunk_id": item.get("chunk_id"),
        "title": item.get("title"),
        "document_type": item.get("document_type"),
        "instructional_unit": item.get("instructional_unit"),
        "source_file": item.get("source_file"),
        "distance": float(item.get("distance")),
    }


def call_experimental_model(openai_module: Any, messages: List[Dict[str, str]]) -> str:
    response = openai_module.ChatCompletion.create(
        model=MODEL,
        messages=messages,
        temperature=TEMPERATURE,
        max_completion_tokens=MAX_COMPLETION_TOKENS,
    )
    message_obj = response.choices[0].message
    return message_obj["content"].strip() if message_obj else ""


def verify_model_compatibility(openai_module: Any) -> Dict[str, Any]:
    try:
        response_text = call_experimental_model(
            openai_module,
            [{"role": "user", "content": "Reply with OK."}],
        )
        return {
            "compatible": True,
            "response": response_text,
            "error": None,
        }
    except Exception as exc:
        return {
            "compatible": False,
            "response": None,
            "error": str(exc),
        }


def run_experiment(cases: List[Dict[str, Any]], v3_by_id: Dict[str, Dict[str, Any]]) -> Dict[str, Any]:
    sys.path.insert(0, str(BACKEND_DIR))
    load_backend_env()

    import chromadb
    import openai
    from chromadb.config import Settings

    from app.core.config import settings
    import app.services.retrieval_service as retrieval_service
    from app.services.llm_service import SYSTEM_PROMPT, build_retrieval_prompt

    api_key = settings.openai_api_key or os.getenv("OPENAI_API_KEY")
    if not api_key:
        raise RuntimeError("Missing OPENAI_API_KEY environment variable")
    openai.api_key = api_key

    compatibility = verify_model_compatibility(openai)
    if not compatibility["compatible"]:
        return {"compatibility": compatibility, "cases": []}

    temp_dir = Path(tempfile.mkdtemp(prefix="ai_mentor_gpt54mini_chroma_"))
    temp_chroma_dir = temp_dir / "chroma_db"
    original_get_chroma_client = retrieval_service.get_chroma_client

    try:
        shutil.copytree(CHROMA_DB_DIR, temp_chroma_dir)
        temp_client = chromadb.Client(
            Settings(persist_directory=str(temp_chroma_dir), is_persistent=True)
        )
        retrieval_service.get_chroma_client = lambda: temp_client

        evaluated = []
        for baseline in cases:
            runtime_error: Optional[str] = None
            answer: Optional[str] = None
            retrieval_context: List[Dict[str, Any]] = []

            try:
                retrieval_context = retrieval_service.retrieve_cpl_context(
                    query=baseline["question"],
                    course_id=COURSE_ID,
                    mentor_mode=baseline["mentor_mode"],
                    top_k=retrieval_service.DEFAULT_TOP_K,
                )
                user_prompt = build_retrieval_prompt(
                    baseline["question"],
                    baseline["mentor_mode"],
                    retrieval_context,
                )
                answer = call_experimental_model(
                    openai,
                    [
                        {"role": "system", "content": SYSTEM_PROMPT},
                        {"role": "user", "content": user_prompt},
                    ],
                )
            except Exception as exc:
                runtime_error = str(exc)

            retrieved_chunks = [
                compact_retrieval_item(item, rank)
                for rank, item in enumerate(retrieval_context, start=1)
            ]
            v3_case = v3_by_id.get(baseline["test_id"], {})
            evaluated.append(
                {
                    "model": MODEL,
                    "test_id": baseline["test_id"],
                    "behaviour_category": baseline["behaviour_category"],
                    "question": baseline["question"],
                    "mentor_mode": baseline["mentor_mode"],
                    "course_id": COURSE_ID,
                    "retrieval_status": "CONTEXT"
                    if retrieved_chunks
                    else "NO_CONTEXT",
                    "retrieved_chunk_count": len(retrieved_chunks),
                    "retrieved_chunk_ids": [
                        item["chunk_id"] for item in retrieved_chunks
                    ],
                    "retrieved_chunk_titles": [
                        item["title"] for item in retrieved_chunks
                    ],
                    "retrieval_distances": [
                        item["distance"] for item in retrieved_chunks
                    ],
                    "retrieved_chunks": retrieved_chunks,
                    "gpt35_v3_response": v3_case.get("prompt_v3_response"),
                    "gpt54mini_response": answer,
                    "api_runtime_error": runtime_error,
                    "llm_response_received": bool(answer) and runtime_error is None,
                }
            )

        return {"compatibility": compatibility, "cases": evaluated}
    finally:
        retrieval_service.get_chroma_client = original_get_chroma_client
        shutil.rmtree(temp_dir, ignore_errors=True)


def summarize(result: Dict[str, Any]) -> Dict[str, Any]:
    cases = result["cases"]
    return {
        "model": MODEL,
        "compatibility_succeeded": result["compatibility"]["compatible"],
        "compatibility_error": result["compatibility"]["error"],
        "tests_executed": len(cases),
        "successful_calls": sum(1 for item in cases if item["llm_response_received"]),
        "api_runtime_errors": sum(1 for item in cases if item["api_runtime_error"]),
        "context_count": sum(
            1 for item in cases if item["retrieval_status"] == "CONTEXT"
        ),
        "no_context_count": sum(
            1 for item in cases if item["retrieval_status"] == "NO_CONTEXT"
        ),
        "target_test_ids": [item["test_id"] for item in cases],
    }


def write_workbook(cases: List[Dict[str, Any]], summary: Dict[str, Any]) -> None:
    workbook = openpyxl.Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Model Summary"
    summary_sheet.append(["Metric", "Value"])
    for key, value in summary.items():
        summary_sheet.append([key, ", ".join(value) if isinstance(value, list) else value])

    case_sheet = workbook.create_sheet("Responses")
    case_sheet.append(
        [
            "Model",
            "Test ID",
            "Behaviour Category",
            "Question",
            "Mentor Mode",
            "Retrieval Status",
            "Retrieved Chunk Count",
            "Retrieved Chunk IDs",
            "Retrieved Chunk Titles",
            "Retrieval Distances",
            "GPT-3.5 V3 Response",
            "GPT-5.4 Mini Response",
            "API Runtime Error",
        ]
    )
    for item in cases:
        case_sheet.append(
            [
                item["model"],
                item["test_id"],
                item["behaviour_category"],
                item["question"],
                item["mentor_mode"],
                item["retrieval_status"],
                item["retrieved_chunk_count"],
                ", ".join(item["retrieved_chunk_ids"]) or "None",
                " | ".join(item["retrieved_chunk_titles"]) or "None",
                ", ".join(f"{distance:.6f}" for distance in item["retrieval_distances"])
                or "None",
                item["gpt35_v3_response"] or "",
                item["gpt54mini_response"] or "",
                item["api_runtime_error"] or "",
            ]
        )

    chunk_sheet = workbook.create_sheet("Retrieved Chunks")
    chunk_sheet.append(
        [
            "Test ID",
            "Rank",
            "Chunk ID",
            "Title",
            "Document Type",
            "Instructional Unit",
            "Distance",
        ]
    )
    for item in cases:
        if not item["retrieved_chunks"]:
            chunk_sheet.append([item["test_id"], "", "NO_CONTEXT", "", "", "", ""])
        for chunk in item["retrieved_chunks"]:
            chunk_sheet.append(
                [
                    item["test_id"],
                    chunk["rank"],
                    chunk["chunk_id"],
                    chunk["title"],
                    chunk["document_type"],
                    chunk["instructional_unit"],
                    round(chunk["distance"], 6),
                ]
            )

    workbook.save(OUTPUT_XLSX)


def write_json(result: Dict[str, Any], summary: Dict[str, Any]) -> None:
    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "stage": "Stage 3 - Model Comparison Experiment",
        "model": MODEL,
        "max_completion_tokens": MAX_COMPLETION_TOKENS,
        "temperature": TEMPERATURE,
        "course_id": COURSE_ID,
        "default_top_k": 5,
        "max_chroma_distance": 1.4,
        "llm_called": True,
        "automated_scoring": False,
        "baseline_gpt35_v3_source": str(BASELINE_V3_JSON),
        "chroma_query_target": "temporary copy of existing ChromaDB contents",
        "compatibility": result["compatibility"],
        "summary": summary,
        "cases": result["cases"],
    }
    OUTPUT_JSON.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def main() -> None:
    cases = load_cases()
    v3_by_id = load_v3_by_id()
    result = run_experiment(cases, v3_by_id)
    summary = summarize(result)

    if not result["compatibility"]["compatible"]:
        print(json.dumps(summary, indent=2))
        raise SystemExit(1)

    write_workbook(result["cases"], summary)
    write_json(result, summary)

    print(f"Wrote {OUTPUT_XLSX}")
    print(f"Wrote {OUTPUT_JSON}")
    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
